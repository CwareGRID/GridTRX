"""
Grid — Browser-based double-entry accounting.
Simple. Analog. Robust.
"""
import os
import sys
import json
import shutil
from datetime import datetime, date
from flask import (Flask, render_template, request, redirect, url_for, 
                   flash, jsonify, send_file, session)
import models

app = Flask(__name__)
app.secret_key = 'grid-accounting-local-use-only'

# ─── Config: where client files live ────────────────────────────────

CONFIG_FILE = None  # Set at startup

def get_config_path():
    """Config file lives next to the program."""
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'grid.json')

def load_config():
    path = get_config_path()
    if os.path.exists(path):
        with open(path, 'r') as f:
            return json.load(f)
    return {'library_path': '', 'last_opened': ''}

def save_config(cfg):
    with open(get_config_path(), 'w') as f:
        json.dump(cfg, f, indent=2)

def get_library_path():
    cfg = load_config()
    return cfg.get('library_path', '')

def list_client_books():
    """Scan the library folder for client book files."""
    lib = get_library_path()
    if not lib or not os.path.isdir(lib):
        return []
    
    clients = []
    for entry in sorted(os.listdir(lib)):
        client_dir = os.path.join(lib, entry)
        if os.path.isdir(client_dir):
            db_path = os.path.join(client_dir, 'books.db')
            exists = os.path.exists(db_path)
            clients.append({
                'name': entry,
                'path': db_path,
                'folder': client_dir,
                'exists': exists,
                'size': os.path.getsize(db_path) if exists else 0,
                'modified': datetime.fromtimestamp(os.path.getmtime(db_path)).strftime('%Y-%m-%d %H:%M') if exists else '',
            })
    return clients

# ─── Jinja2 Filters ────────────────────────────────────────────────

@app.before_request
def load_company_info():
    from flask import g
    if models.get_db_path():
        g.company_name = models.get_meta('company_name', '')
        fye = models.get_meta('fiscal_year_end', '')  # MM-DD
        fy_year = models.get_meta('fiscal_year', '')   # YYYY
        # Format as "31 Dec 2025"
        if fye and fy_year:
            months = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
                      7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
            try:
                parts = fye.split('-')
                mm, dd = int(parts[0]), int(parts[1])
                g.fiscal_display = f"{dd} {months.get(mm,'?')} {fy_year}"
            except:
                g.fiscal_display = f"{fye} {fy_year}"
        elif fye:
            g.fiscal_display = fye
        else:
            g.fiscal_display = ''

@app.template_filter('money')
def money_filter(cents):
    return models.fmt_amount(cents)

@app.template_filter('money_plain')
def money_plain_filter(cents):
    return models.fmt_amount_plain(cents)

@app.template_filter('money_dr')
def money_dr_filter(cents):
    return models.fmt_amount(cents) if cents > 0 else ''

@app.template_filter('money_cr')
def money_cr_filter(cents):
    return models.fmt_amount(abs(cents)) if cents < 0 else ''

# Make fmt_amount available as a global in templates
app.jinja_env.globals['fmt'] = models.fmt_amount

def pct_fmt(basis_points):
    """Format basis points as percentage string. 10000 bp = 100.0%"""
    if basis_points is None or basis_points == 0:
        return '—'
    pct = basis_points / 100.0
    return f'{pct:,.1f}%'

app.jinja_env.globals['pct_fmt'] = pct_fmt

# ─── File Picker / Library ──────────────────────────────────────────

@app.route('/library')
def library():
    """The starting screen — pick a client or create a new one."""
    cfg = load_config()
    lib_path = cfg.get('library_path', '')
    clients = list_client_books() if lib_path else []
    return render_template('library.html', clients=clients, library_path=lib_path)

@app.route('/library/set-path', methods=['POST'])
def set_library_path():
    """Set the folder where client files live."""
    path = request.form.get('library_path', '').strip()
    if path and os.path.isdir(path):
        cfg = load_config()
        cfg['library_path'] = path
        save_config(cfg)
        flash(f'Library path set to: {path}', 'success')
    else:
        flash('That folder does not exist. Please check the path.', 'error')
    return redirect(url_for('library'))

@app.route('/library/new-client', methods=['POST'])
def new_client():
    """Create a new client folder with empty books."""
    name = request.form.get('client_name', '').strip()
    if not name:
        flash('Please enter a client name.', 'error')
        return redirect(url_for('library'))
    
    lib = get_library_path()
    if not lib:
        flash('Set your library path first.', 'error')
        return redirect(url_for('library'))
    
    client_dir = os.path.join(lib, name)
    db_path = os.path.join(client_dir, 'books.db')
    
    if os.path.exists(client_dir):
        flash(f'A folder named "{name}" already exists.', 'error')
        return redirect(url_for('library'))
    
    os.makedirs(client_dir)
    models.create_starter_books(db_path, name)
    
    # Save as last opened
    cfg = load_config()
    cfg['last_opened'] = db_path
    save_config(cfg)
    
    flash(f'Created new books for {name}', 'success')
    return redirect(url_for('open_client', path=db_path))

@app.route('/open')
def open_client():
    """Open a specific client's books."""
    path = request.args.get('path', '')
    if not path or not os.path.exists(path):
        flash('File not found.', 'error')
        return redirect(url_for('library'))
    
    models.set_db_path(path)
    models._ensure_columns()  # Migrate older DBs
    cfg = load_config()
    cfg['last_opened'] = path
    save_config(cfg)
    
    return redirect(url_for('home'))

@app.route('/close')
def close_books():
    """Close current books and return to the library/startup screen."""
    models.set_db_path(None)
    cfg = load_config()
    cfg['last_opened'] = ''
    save_config(cfg)
    return redirect(url_for('library'))

# ─── Home ───────────────────────────────────────────────────────────

@app.route('/')
def home():
    if not models.get_db_path():
        return redirect(url_for('library'))
    reports = models.get_reports()
    company = models.get_meta('company_name', 'My Books')
    db_path = models.get_db_path()
    client_folder = os.path.basename(os.path.dirname(db_path))
    return render_template('home.html', reports=reports, company=company,
                         client_folder=client_folder)

# ─── Report View (with total-to chains) ───────────────────────────

@app.route('/report/<int:report_id>')
def report_view(report_id):
    import time as _time
    _t0 = _time.time()
    
    if not models.get_db_path():
        return redirect(url_for('library'))
    report = models.get_report(report_id)
    if not report:
        flash('Report not found', 'error')
        return redirect(url_for('home'))
    
    company = models.get_meta('company_name', 'My Books')
    hide_zero = request.args.get('hide_zero', '0') == '1'
    show_setup = request.args.get('show_setup', '0') == '1'
    
    # ── Column Config: persist per report ──
    import json as _json
    columns = []
    items = models.get_report_items(report_id)
    all_items = models.get_all_report_items()  # fetch once, reuse
    
    # Check if user submitted new config via Apply button
    from_query = any(request.args.get(f'c{i}_end') or request.args.get(f'c{i}_type') in ('change','pct_change','spacer')
                     for i in range(1, 7))
    
    if from_query:
        # Save to meta
        cfg = {}
        for i in range(1, 7):
            for k in ('begin', 'end', 'type', 'label', 'a', 'b'):
                v = request.args.get(f'c{i}_{k}', '')
                if v: cfg[f'c{i}_{k}'] = v
        models.set_meta(f'columns_{report_id}', _json.dumps(cfg))
    elif request.args.get('reset'):
        cfg = {}
        models.set_meta(f'columns_{report_id}', '')
    else:
        # Load from meta
        raw = models.get_meta(f'columns_{report_id}', '')
        cfg = _json.loads(raw) if raw else {}
    
    def carg(k, d=''):
        return cfg.get(k, d)
    
    has_custom = any(carg(f'c{i}_end') or carg(f'c{i}_type') in ('change','pct_change','spacer') for i in range(1, 7))
    
    # Build all 6 column slots (None for unused) to keep positions stable
    all_columns = [None] * 6
    if has_custom:
        for i in range(1, 7):
            end = carg(f'c{i}_end', '')
            begin = carg(f'c{i}_begin', '')
            ctype = carg(f'c{i}_type', 'A')
            label = carg(f'c{i}_label', '')
            if not end and ctype not in ('change', 'pct_change', 'spacer'):
                continue
            if ctype == 'change':
                ca = int(carg(f'c{i}_a', '1')) - 1
                cb = int(carg(f'c{i}_b', '2')) - 1
                all_columns[i-1] = {'type': 'change', 'a': ca, 'b': cb,
                               'label': label or '$ chg', 'data': None}
            elif ctype == 'pct_change':
                ca = int(carg(f'c{i}_a', '1')) - 1
                cb = int(carg(f'c{i}_b', '2')) - 1
                all_columns[i-1] = {'type': 'pct_change', 'a': ca, 'b': cb,
                               'label': label or '% chg', 'data': None}
            elif ctype == 'spacer':
                all_columns[i-1] = {'type': 'spacer', 'label': label or '', 'data': None}
            else:
                col_data = models.compute_report_column(report_id,
                    date_from=begin or None, date_to=end or None,
                    _display_items=items, _all_items=all_items)
                all_columns[i-1] = {'type': 'actual', 'begin': begin, 'end': end,
                               'label': label or end[:4] if end else 'Current',
                               'data': col_data}
        # Active columns in display order (left to right, matching config 1→6)
        columns = [c for c in all_columns if c is not None]
    
    if not columns:
        # Default: single column, all dates
        col_data = models.compute_report_column(report_id,
            _display_items=items, _all_items=all_items)
        columns.append({'type': 'actual', 'begin': '', 'end': '',
                       'label': 'Balance', 'data': col_data})
        all_columns = [None] * 6
    
    # For change/pct columns, compute deltas; for spacers, fill with None
    for col in columns:
        if col['type'] == 'change':
            a_idx, b_idx = col['a'], col['b']
            if a_idx < len(columns) and b_idx < len(columns):
                a_data = columns[a_idx].get('data', [])
                b_data = columns[b_idx].get('data', [])
                if a_data and b_data:
                    change_data = []
                    for j in range(len(a_data)):
                        item_a, bal_a = a_data[j]
                        _, bal_b = b_data[j] if j < len(b_data) else (None, 0)
                        change_data.append((item_a, bal_b - bal_a))
                    col['data'] = change_data
        elif col['type'] == 'pct_change':
            a_idx, b_idx = col['a'], col['b']
            if a_idx < len(columns) and b_idx < len(columns):
                a_data = columns[a_idx].get('data', [])
                b_data = columns[b_idx].get('data', [])
                if a_data and b_data:
                    pct_data = []
                    for j in range(len(a_data)):
                        item_a, bal_a = a_data[j]
                        _, bal_b = b_data[j] if j < len(b_data) else (None, 0)
                        if bal_a != 0:
                            pct = round((bal_b - bal_a) * 10000 / abs(bal_a))  # basis points → will format later
                        else:
                            pct = 0  # avoid div/0
                        pct_data.append((item_a, pct))
                    col['data'] = pct_data
        elif col['type'] == 'spacer':
            # Spacer needs data array matching row count for iteration
            pass  # handled below
    
    # Build unified row structure: items + balance per column
    rows = []
    base_items = columns[0]['data'] if columns[0].get('data') else []
    # If first column is a spacer, find first actual data column
    if not base_items:
        for col in columns:
            if col.get('data'):
                base_items = col['data']
                break
    for idx, (item, _) in enumerate(base_items):
        bals = []
        for col in columns:
            if col['type'] == 'spacer':
                bals.append(None)  # None = spacer cell
            elif col.get('data') and idx < len(col['data']):
                bals.append(col['data'][idx][1])
            else:
                bals.append(0)
        rows.append((item, bals))
    
    col_labels = [c['label'] for c in columns]
    col_types = [c['type'] for c in columns]
    
    _elapsed = (_time.time() - _t0) * 1000
    app.logger.info(f'Report {report_id} data computed in {_elapsed:.0f}ms')
    
    return render_template('report.html', report=report, rows=rows,
                         col_labels=col_labels, col_types=col_types, 
                         columns=columns, all_columns=all_columns, cfg=cfg,
                         company=company, hide_zero=hide_zero, show_setup=show_setup)

@app.route('/report/<int:report_id>/print')
def report_print(report_id):
    """Print-friendly report view. Options: begin, end, hide_zero, ledger, debit, credit."""
    report = models.get_report(report_id)
    if not report:
        return 'Report not found', 404
    company = models.get_meta('company_name', 'My Books')
    begin = request.args.get('begin', '')
    end = request.args.get('end', '')
    hide_zero = request.args.get('hide_zero', '0') == '1'
    mode = request.args.get('mode', 'report')  # report, ledger, debit, credit

    if mode == 'ledger':
        # Full ledger dump for all posting accounts in this report
        items = models.get_report_items(report_id)
        ledger_data = []
        for it in items:
            if it['account_id'] and it['account_type'] == 'posting':
                entries = models.get_ledger(it['account_id'],
                    date_from=begin or None, date_to=end or None)
                if entries or not hide_zero:
                    ledger_data.append({
                        'name': it['acct_name'],
                        'desc': it['description'] or it['acct_desc'] or it['acct_name'],
                        'entries': entries
                    })
        return render_template('print_ledger.html', report=report, company=company,
                             ledger_data=ledger_data, begin=begin, end=end, mode=mode,
                             now=__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M'))
    else:
        # Check for saved multi-column config
        raw = models.get_meta(f'columns_{report_id}', '')
        cfg = json.loads(raw) if raw else {}
        def carg(k, d=''):
            return cfg.get(k, d)

        has_custom = any(carg(f'c{i}_end') or carg(f'c{i}_type') in ('change','pct_change','spacer') for i in range(1, 7))

        if has_custom:
            # Multi-column print: replicate the on-screen column logic
            items = models.get_report_items(report_id)
            all_items = models.get_all_report_items()

            all_columns = [None] * 6
            for i in range(1, 7):
                end_c = carg(f'c{i}_end', '')
                begin_c = carg(f'c{i}_begin', '')
                ctype = carg(f'c{i}_type', 'A')
                label = carg(f'c{i}_label', '')
                if not end_c and ctype not in ('change', 'pct_change', 'spacer'):
                    continue
                if ctype == 'change':
                    ca = int(carg(f'c{i}_a', '1')) - 1
                    cb = int(carg(f'c{i}_b', '2')) - 1
                    all_columns[i-1] = {'type': 'change', 'a': ca, 'b': cb,
                                   'label': label or '$ chg', 'data': None}
                elif ctype == 'pct_change':
                    ca = int(carg(f'c{i}_a', '1')) - 1
                    cb = int(carg(f'c{i}_b', '2')) - 1
                    all_columns[i-1] = {'type': 'pct_change', 'a': ca, 'b': cb,
                                   'label': label or '% chg', 'data': None}
                elif ctype == 'spacer':
                    all_columns[i-1] = {'type': 'spacer', 'label': label or '', 'data': None}
                else:
                    col_data = models.compute_report_column(report_id,
                        date_from=begin_c or None, date_to=end_c or None,
                        _display_items=items, _all_items=all_items)
                    all_columns[i-1] = {'type': 'actual', 'begin': begin_c, 'end': end_c,
                                   'label': label or end_c[:4] if end_c else 'Current',
                                   'data': col_data}

            columns = [c for c in all_columns if c is not None]

            if not columns:
                # Fallback to single column
                col_data = models.compute_report_column(report_id,
                    _display_items=items, _all_items=all_items)
                columns = [{'type': 'actual', 'begin': '', 'end': '',
                           'label': 'Balance', 'data': col_data}]

            # Compute change/pct columns
            for col in columns:
                if col['type'] == 'change':
                    a_idx, b_idx = col['a'], col['b']
                    if a_idx < len(columns) and b_idx < len(columns):
                        a_data = columns[a_idx].get('data', [])
                        b_data = columns[b_idx].get('data', [])
                        if a_data and b_data:
                            change_data = []
                            for j in range(len(a_data)):
                                item_a, bal_a = a_data[j]
                                _, bal_b = b_data[j] if j < len(b_data) else (None, 0)
                                change_data.append((item_a, bal_b - bal_a))
                            col['data'] = change_data
                elif col['type'] == 'pct_change':
                    a_idx, b_idx = col['a'], col['b']
                    if a_idx < len(columns) and b_idx < len(columns):
                        a_data = columns[a_idx].get('data', [])
                        b_data = columns[b_idx].get('data', [])
                        if a_data and b_data:
                            pct_data = []
                            for j in range(len(a_data)):
                                item_a, bal_a = a_data[j]
                                _, bal_b = b_data[j] if j < len(b_data) else (None, 0)
                                if bal_a != 0:
                                    pct = round((bal_b - bal_a) * 10000 / abs(bal_a))
                                else:
                                    pct = 0
                                pct_data.append((item_a, pct))
                            col['data'] = pct_data

            # Build rows
            base_items = columns[0]['data'] if columns[0].get('data') else []
            if not base_items:
                for col in columns:
                    if col.get('data'):
                        base_items = col['data']
                        break
            rows = []
            for idx, (item, _) in enumerate(base_items):
                bals = []
                for col in columns:
                    if col['type'] == 'spacer':
                        bals.append(None)
                    elif col.get('data') and idx < len(col['data']):
                        bals.append(col['data'][idx][1])
                    else:
                        bals.append(0)
                if hide_zero and item.get('item_type') in ('account',) and all((b is None or b == 0) for b in bals):
                    continue
                rows.append((item, bals))

            col_labels = [c['label'] for c in columns]
            col_types = [c['type'] for c in columns]

            return render_template('print_report.html', report=report, company=company,
                                 rows=rows, begin='', end='', hide_zero=hide_zero,
                                 col_labels=col_labels, col_types=col_types, multicol=True,
                                 now=__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M'))
        else:
            # Single column print (original behavior)
            col_data = models.compute_report_column(report_id,
                date_from=begin or None, date_to=end or None)
            rows = []
            for item, bal in col_data:
                if hide_zero and bal == 0 and item.get('item_type') in ('account',):
                    continue
                rows.append((item, bal))
            return render_template('print_report.html', report=report, company=company,
                                 rows=rows, begin=begin, end=end, hide_zero=hide_zero,
                                 col_labels=['Balance'], col_types=['actual'], multicol=False,
                                 now=__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M'))

@app.route('/report/<int:report_id>/csv')
def report_csv(report_id):
    """Export report as CSV."""
    import csv, io
    report = models.get_report(report_id)
    if not report:
        return 'Not found', 404
    begin = request.args.get('begin', '')
    end = request.args.get('end', '')
    col_data = models.compute_report_column(report_id,
        date_from=begin or None, date_to=end or None)
    
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['Description', 'Account', 'Type', 'Balance'])
    for item, bal in col_data:
        desc = item.get('description') or item.get('acct_desc') or item.get('acct_name', '')
        name = item.get('acct_name', '')
        itype = item.get('item_type', '')
        if itype in ('label', 'separator'):
            w.writerow([desc, '', itype, ''])
        else:
            w.writerow([desc, name, itype, f'{bal/100:.2f}' if bal else ''])
    
    from flask import Response
    return Response(buf.getvalue(), mimetype='text/csv',
                   headers={'Content-Disposition': f'attachment; filename={report["name"]}.csv'})

# ─── Multi-Column (13-Column) Report ─────────────────────────────

@app.route('/report/<int:report_id>/multicol')
def report_multicol(report_id):
    """Generate a multi-column monthly report as landscape PDF.
    Params: start (yyyy-mm), months (1-24), mode (periodic|cumulative), hide_zero."""
    from calendar import monthrange
    from datetime import date as _date
    
    report = models.get_report(report_id)
    if not report:
        return 'Report not found', 404
    company = models.get_meta('company_name', 'My Books')
    
    start = request.args.get('start', '')
    num_months = int(request.args.get('months', '12'))
    mode = request.args.get('mode', 'periodic')  # periodic or cumulative
    hide_zero = request.args.get('hide_zero', '0') == '1'
    
    if not start:
        # Show config form
        fye = models.get_meta('fiscal_year_end', '12-31')
        fy_year = models.get_meta('fiscal_year', str(_date.today().year))
        # Default start: beginning of fiscal year
        fye_mm = fye.split('-')[0] if '-' in fye else '01'
        default_start = f'{fy_year}-{fye_mm}'
        return f'''<!DOCTYPE html><html><head><title>Multi-Column Report</title>
        <style>body{{font-family:system-ui;max-width:500px;margin:40px auto;background:#1a1a2e;color:#e0e0e0}}
        label{{display:block;margin:10px 0 4px;font-size:13px;font-weight:600}}
        input,select{{padding:6px 10px;font-size:14px;border:1px solid #444;border-radius:4px;
        background:#2a2a4a;color:#e0e0e0;font-family:monospace}}
        .btn{{padding:8px 20px;background:#4a6fa5;color:white;border:none;border-radius:4px;
        cursor:pointer;font-size:14px;margin-top:14px}}
        h2{{color:#8ab4f8}}</style></head><body>
        <h2>{report["name"]} — Multi-Column Report</h2>
        <form>
        <label>Start Month (yyyy-mm)</label>
        <input name="start" value="{default_start}" placeholder="2025-01">
        <label>Number of Months</label>
        <input name="months" value="12" type="number" min="1" max="36" style="width:60px">
        <label>Mode</label>
        <select name="mode">
        <option value="periodic">Periodic (each month standalone — use for Income Statement)</option>
        <option value="cumulative">Cumulative (running balance — use for Balance Sheet)</option>
        </select>
        <label><input type="checkbox" name="hide_zero" value="1" style="width:auto"> Hide zero rows</label>
        <br><button type="submit" class="btn">Generate PDF</button>
        </form></body></html>'''
    
    # Build month ranges
    import re
    m = re.match(r'^(\d{4})-(\d{2})$', start)
    if not m:
        return 'Invalid start format. Use yyyy-mm', 400
    
    start_year, start_month = int(m.group(1)), int(m.group(2))
    
    month_ranges = []
    for i in range(num_months):
        y = start_year + (start_month - 1 + i) // 12
        mo = (start_month - 1 + i) % 12 + 1
        last_day = monthrange(y, mo)[1]
        d_from = f'{y:04d}-{mo:02d}-01'
        d_to = f'{y:04d}-{mo:02d}-{last_day:02d}'
        
        if mode == 'cumulative':
            # BS mode: from the beginning of time to end of this month
            month_ranges.append((None, d_to, f'{y:04d}-{mo:02d}'))
        else:
            # IS mode: just this month
            month_ranges.append((d_from, d_to, f'{y:04d}-{mo:02d}'))
    
    # Compute columns — prefetch items once
    items = models.get_report_items(report_id)
    all_items = models.get_all_report_items()
    
    columns = []
    for d_from, d_to, label in month_ranges:
        col_data = models.compute_report_column(report_id,
            date_from=d_from, date_to=d_to,
            _display_items=items, _all_items=all_items)
        columns.append((label, col_data))
    
    # Build rows: use first column for structure
    if not columns or not columns[0][1]:
        return 'No data', 404
    
    base = columns[0][1]
    row_data = []
    for idx, (item, _) in enumerate(base):
        vals = []
        for _, col in columns:
            if idx < len(col):
                vals.append(col[idx][1])
            else:
                vals.append(0)
        # Total column
        if mode == 'cumulative':
            total = vals[-1] if vals else 0  # last month IS the cumulative total
        else:
            total = sum(vals)
        
        if hide_zero and item.get('item_type') == 'account' and all(v == 0 for v in vals):
            continue
        row_data.append((item, vals, total))
    
    # Generate landscape PDF
    pdf_bytes = _multicol_pdf(report, company, columns, row_data, mode, num_months)
    
    from flask import Response
    resp = Response(pdf_bytes, mimetype='application/pdf')
    resp.headers['Content-Disposition'] = f'inline; filename="{report["name"]}_multicol.pdf"'
    return resp


def _multicol_pdf(report, company, columns, row_data, mode, num_months):
    """Generate a landscape multi-column PDF report."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        from flask import abort
        abort(500, 'reportlab not installed')
    import io, os
    
    pw, ph = letter[1], letter[0]  # landscape
    margin = 36  # 0.5"
    
    # Font setup (same as GL PDF)
    font = 'Courier'
    font_b = 'Courier-Bold'
    candidates = [
        ('LiberationMono', '/usr/share/fonts/truetype/liberation/LiberationMono-Regular.ttf',
                           '/usr/share/fonts/truetype/liberation/LiberationMono-Bold.ttf'),
        ('DejaVuMono',     '/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf',
                           '/usr/share/fonts/truetype/dejavu/DejaVuSansMono-Bold.ttf'),
        ('Consolas',       'C:/Windows/Fonts/consola.ttf',
                           'C:/Windows/Fonts/consolab.ttf'),
        ('CourierNew',     'C:/Windows/Fonts/cour.ttf',
                           'C:/Windows/Fonts/courbd.ttf'),
        ('Menlo',          '/System/Library/Fonts/Menlo.ttc',
                           '/System/Library/Fonts/Menlo.ttc'),
    ]
    for fname, reg_path, bold_path in candidates:
        if os.path.exists(reg_path):
            try:
                pdfmetrics.registerFont(TTFont(fname, reg_path))
                pdfmetrics.registerFont(TTFont(fname + '-Bold', bold_path))
                font = fname
                font_b = fname + '-Bold'
            except:
                pass
            break
    
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(pw, ph))
    
    ncols = len(columns)
    show_total = (mode == 'periodic')  # periodic gets a total column; cumulative doesn't need one
    num_data_cols = ncols + (1 if show_total else 0)
    
    # Layout calculation
    usable = pw - 2 * margin
    # Account description gets leftover space after amount columns
    # Each amount column needs ~58pt for numbers like (999,999)
    col_w = 56
    if num_data_cols > 13:
        col_w = 50  # squeeze for >13 columns
    if num_data_cols > 18:
        col_w = 44
    
    desc_w = usable - (num_data_cols * col_w)
    if desc_w < 100:
        desc_w = 100
        col_w = max(40, (usable - desc_w) // num_data_cols)
    
    fs = 6.5
    if num_data_cols > 14:
        fs = 6
    if num_data_cols > 20:
        fs = 5.5
    line_h = fs + 2.5
    
    col_starts = []
    x = margin + desc_w
    for i in range(num_data_cols):
        col_starts.append(x)
        x += col_w
    
    y = ph - margin
    page_num = 1
    right_edge = pw - margin
    
    mode_label = 'Cumulative' if mode == 'cumulative' else 'Periodic'
    first_label = columns[0][0] if columns else ''
    last_label = columns[-1][0] if columns else ''
    
    def header():
        nonlocal y
        c.setFont(font_b, 8)
        c.drawString(margin, ph - margin + 5, f'{company} — {report["name"]} ({mode_label})')
        c.setFont(font, 6)
        c.drawString(margin, ph - margin - 4, f'{first_label} to {last_label}')
        c.drawRightString(right_edge, ph - margin + 5, f'Page {page_num}')
        y = ph - margin - 14
    
    def col_header():
        nonlocal y
        c.setFont(font_b, fs - 0.5)
        c.drawString(margin, y, 'Account')
        for i, (label, _) in enumerate(columns):
            # Show month abbreviation: yyyy-mm -> Mon
            try:
                yr, mo = label.split('-')
                from calendar import month_abbr
                short = month_abbr[int(mo)]
                if ncols <= 12:
                    short = f'{short} {yr[2:]}'
                c.drawRightString(col_starts[i] + col_w - 2, y, short)
            except:
                c.drawRightString(col_starts[i] + col_w - 2, y, label)
        if show_total:
            c.drawRightString(col_starts[-1] + col_w - 2, y, 'Total')
        y -= 2
        c.setLineWidth(0.4)
        c.line(margin, y, right_edge, y)
        y -= line_h
    
    def check_page():
        nonlocal y, page_num
        if y < margin + 2 * line_h:
            c.showPage()
            page_num += 1
            header()
            col_header()
    
    def fmt_val(v):
        """Format cents as string with parens for negative, em dash for zero."""
        if v == 0:
            return '\u2014'
        if v < 0:
            return f'({abs(v)/100:,.0f})'
        return f'{v/100:,.0f}'
    
    header()
    col_header()
    
    for item, vals, total in row_data:
        check_page()
        itype = item.get('item_type', 'account')
        indent = item.get('indent', 0) or 0
        
        if itype == 'separator':
            style = item.get('sep_style', 'single')
            if style == 'double':
                c.setLineWidth(0.5)
                c.line(col_starts[0], y + line_h * 0.4, right_edge, y + line_h * 0.4)
                c.line(col_starts[0], y + line_h * 0.4 - 2, right_edge, y + line_h * 0.4 - 2)
            elif style == 'blank':
                pass
            else:
                c.setLineWidth(0.3)
                c.line(col_starts[0], y + line_h * 0.4, right_edge, y + line_h * 0.4)
            y -= line_h
            continue
        
        desc = item.get('description') or item.get('acct_desc') or item.get('acct_name') or ''
        is_total = itype == 'total'
        fn = font_b if is_total else font
        
        c.setFont(fn, fs)
        # Truncate description to fit
        max_desc_chars = int(desc_w / (fs * 0.6))
        display_desc = '  ' * indent + desc
        c.drawString(margin, y, display_desc[:max_desc_chars])
        
        # Draw month values
        for i, v in enumerate(vals):
            c.drawRightString(col_starts[i] + col_w - 2, y, fmt_val(v))

        # Total column
        if show_total:
            c.setFont(font_b, fs)
            c.drawRightString(col_starts[-1] + col_w - 2, y, fmt_val(total))
        
        y -= line_h
    
    c.save()
    return buf.getvalue()


# ─── Report Item Management API ──────────────────────────────────

@app.route('/api/report/<int:report_id>/add-item', methods=['POST'])
def api_add_report_item(report_id):
    """Add a new account/item to a report."""
    try:
        item_type = request.form.get('item_type', 'account')
        acct_name = request.form.get('account_name', '').strip().upper()
        description = request.form.get('description', '')
        total_to_1 = request.form.get('total_to_1', '').strip().upper()
        after_pos = int(request.form.get('after_position', '0'))
        indent = int(request.form.get('indent', '2'))
        sep_style = request.form.get('sep_style', '')
        nb = request.form.get('normal_balance', 'D')
        
        account_id = None
        if item_type in ('account', 'total') and acct_name:
            acct = models.get_account_by_name(acct_name)
            if not acct:
                atype = 'total' if item_type == 'total' else 'posting'
                account_id = models.add_account(acct_name, nb, description, atype)
            else:
                account_id = acct['id']
                # Update the existing account's normal_balance and description
                # in case the user is correcting a previous mistake
                with models.get_db() as db:
                    db.execute("UPDATE accounts SET normal_balance=? WHERE id=?", (nb, acct['id']))
                    if description:
                        db.execute("UPDATE accounts SET description=? WHERE id=?", (description, acct['id']))
        
        position = after_pos + 5 if after_pos else None
        
        models.add_report_item(report_id, item_type, description, account_id,
                              indent, position, total_to_1, sep_style=sep_style)
        
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/report-item/<int:item_id>/update', methods=['POST'])
def api_update_report_item(item_id):
    """Update fields on a report item (total-to's, description, indent, sep_style)."""
    try:
        data = request.get_json() if request.is_json else request.form
        kwargs = {}
        for field in ('description','indent','total_to_1','total_to_2','total_to_3',
                      'total_to_4','total_to_5','total_to_6','sep_style','position','item_type'):
            if field in data:
                val = data[field]
                if field == 'indent':
                    val = int(val)
                elif field == 'position':
                    val = int(val)
                elif field.startswith('total_to_'):
                    val = str(val).strip().upper()
                kwargs[field] = val
        models.update_report_item(item_id, **kwargs)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/report-item/<int:item_id>/delete', methods=['POST'])
def api_delete_report_item(item_id):
    """Delete a report item (with safety checks)."""
    try:
        models.delete_report_item(item_id)
        return jsonify({'ok': True})
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/report-item/<int:item_id>/move', methods=['POST'])
def api_move_report_item(item_id):
    """Move a report item up or down."""
    try:
        data = request.get_json() if request.is_json else request.form
        direction = int(data.get('direction', 0))
        if direction not in (-1, 1):
            return jsonify({'ok': False, 'error': 'Direction must be -1 or 1'})
        moved = models.move_report_item(item_id, direction)
        return jsonify({'ok': True, 'moved': moved})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/account/<int:account_id>/rename', methods=['POST'])
def api_rename_account(account_id):
    """Rename an account's description (not the code name)."""
    try:
        data = request.get_json() if request.is_json else request.form
        desc = data.get('description', '')
        acct_num = data.get('account_number')
        models.update_account(account_id, description=desc,
            account_number=acct_num if acct_num is not None else None)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/reports/reorder', methods=['POST'])
def api_reorder_reports():
    """Reorder reports. Expects JSON: {order: [id, id, id, ...]}"""
    try:
        data = request.get_json()
        order = data.get('order', [])
        with models.get_db() as db:
            for i, rid in enumerate(order):
                db.execute("UPDATE reports SET sort_order=? WHERE id=?", (i * 10, rid))
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/report/<int:report_id>/sort', methods=['POST'])
def api_sort_report(report_id):
    """Set a single report's sort order."""
    try:
        data = request.get_json()
        sort_order = int(data.get('sort_order', 0))
        with models.get_db() as db:
            db.execute("UPDATE reports SET sort_order=? WHERE id=?", (sort_order, report_id))
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# ─── Account Ledger ─────────────────────────────────────────────────

@app.route('/account/<int:account_id>')
@app.route('/ledger/<int:account_id>')
def account_ledger(account_id):
    if not models.get_db_path():
        return redirect(url_for('library'))
    account = models.get_account(account_id)
    if not account:
        flash('Account not found', 'error')
        return redirect(url_for('home'))
    
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')
    
    entries = models.get_ledger(account_id, date_from or None, date_to or None)
    balance = models.get_account_balance(account_id)
    sign = 1 if account['normal_balance'] == 'D' else -1
    balance = balance * sign
    company = models.get_meta('company_name', 'My Books')
    
    # Track parent report for breadcrumb navigation
    # If from_report specified, use it; otherwise auto-detect from BS or IS
    from_report_id = request.args.get('from_report', '')
    parent_report = None
    if from_report_id:
        parent_report = models.get_report(int(from_report_id))
    if not parent_report:
        parent_report = models.find_report_for_account(account_id)
    
    return render_template('ledger.html', account=account, entries=entries,
                         balance=balance, date_from=date_from, date_to=date_to,
                         company=company, today=date.today().isoformat(),
                         parent_report=parent_report)

# ─── Cross-Account Jump ────────────────────────────────────────────

@app.route('/ledger-by-name/<name>')
def ledger_by_name(name):
    """Jump to an account's ledger by name. Used from F5 in distribution view."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    acct = models.get_account_by_name(name.upper())
    if not acct:
        flash(f'Account "{name}" not found', 'error')
        return redirect(url_for('home'))
    focus_txn = request.args.get('focus_txn', '')
    from_report = request.args.get('from_report', '')
    url = f'/ledger/{acct["id"]}'
    params = []
    if from_report: params.append(f'from_report={from_report}')
    if params: url += '?' + '&'.join(params)
    if focus_txn: url += f'#txn-{focus_txn}'
    return redirect(url)

@app.route('/jump/<int:txn_id>/from/<int:from_account_id>')
def jump_to_cross(txn_id, from_account_id):
    """Jump to the cross-account of a transaction."""
    txn, lines = models.get_transaction(txn_id)
    if not txn:
        flash('Transaction not found', 'error')
        return redirect(url_for('home'))
    
    for line in lines:
        if line['account_id'] != from_account_id:
            return redirect(url_for('account_ledger', account_id=line['account_id'],
                                  _anchor=f'txn-{txn_id}'))
    return redirect(url_for('account_ledger', account_id=from_account_id))

# ─── New Transaction ────────────────────────────────────────────────

@app.route('/transaction/new', methods=['GET', 'POST'])
@app.route('/transaction/new/in/<int:account_id>', methods=['GET', 'POST'])
def new_transaction(account_id=None):
    if not models.get_db_path():
        return redirect(url_for('library'))
    
    if request.method == 'POST':
        try:
            date_str = request.form['date']
            reference = request.form.get('reference', '')
            description = request.form.get('description', '')
            mode = request.form.get('mode', 'simple')
            
            if mode == 'simple':
                debit_name = request.form['debit_account']
                credit_name = request.form['credit_account']
                amount_str = request.form['amount']
                
                debit_acct = models.get_account_by_name(debit_name)
                credit_acct = models.get_account_by_name(credit_name)
                
                if not debit_acct:
                    flash(f'Account not found: {debit_name}', 'error')
                    return redirect(request.url)
                if not credit_acct:
                    flash(f'Account not found: {credit_name}', 'error')
                    return redirect(request.url)
                
                amount = models.parse_amount(amount_str)
                txn_id = models.add_simple_transaction(
                    date_str, reference, description,
                    debit_acct['id'], credit_acct['id'], amount)
            else:
                lines = []
                acct_names = request.form.getlist('line_account[]')
                amounts = request.form.getlist('line_amount[]')
                descs = request.form.getlist('line_desc[]')
                
                for acct_name, amt_str, desc in zip(acct_names, amounts, descs):
                    if not acct_name.strip() or not amt_str.strip():
                        continue
                    acct = models.get_account_by_name(acct_name.strip())
                    if not acct:
                        flash(f'Account not found: {acct_name}', 'error')
                        return redirect(request.url)
                    amount = models.parse_amount(amt_str)
                    lines.append((acct['id'], amount, desc))
                
                txn_id = models.add_transaction(date_str, reference, description, lines)
            
            flash('Transaction posted', 'success')
            if account_id:
                return redirect(url_for('account_ledger', account_id=account_id))
            return redirect(url_for('home'))
            
        except ValueError as e:
            flash(str(e), 'error')
            return redirect(request.url)
    
    account = models.get_account(account_id) if account_id else None
    accounts = models.get_accounts()
    today = date.today().isoformat()
    company = models.get_meta('company_name', 'My Books')
    return render_template('transaction.html', account=account, accounts=accounts,
                         today=today, company=company)

# ─── Edit Transaction ───────────────────────────────────────────────

@app.route('/transaction/<int:txn_id>/edit', methods=['GET', 'POST'])
def edit_transaction(txn_id):
    if not models.get_db_path():
        return redirect(url_for('library'))
    
    if request.method == 'POST':
        try:
            date_str = request.form['date']
            reference = request.form.get('reference', '')
            description = request.form.get('description', '')
            return_to = request.form.get('return_to', '')
            
            lines = []
            acct_names = request.form.getlist('line_account[]')
            amounts = request.form.getlist('line_amount[]')
            descs = request.form.getlist('line_desc[]')
            reconcileds = request.form.getlist('line_reconciled[]')
            doc_flags = request.form.getlist('line_doc_on_file[]')
            
            for i, (acct_name, amt_str, desc) in enumerate(zip(acct_names, amounts, descs)):
                if not acct_name.strip() or not amt_str.strip():
                    continue
                acct = models.get_account_by_name(acct_name.strip())
                if not acct:
                    flash(f'Account not found: {acct_name}', 'error')
                    return redirect(request.url)
                amount = models.parse_amount(amt_str)
                rec = int(reconcileds[i]) if i < len(reconcileds) else 0
                doc = int(doc_flags[i]) if i < len(doc_flags) else 0
                lines.append((acct['id'], amount, desc, rec, doc))
            
            models.update_transaction(txn_id, date_str, reference, description, lines)
            flash('Transaction updated', 'success')
            
            if return_to:
                return redirect(return_to)
            return redirect(url_for('home'))
            
        except ValueError as e:
            flash(str(e), 'error')
            return redirect(request.url)
    
    txn, lines = models.get_transaction(txn_id)
    if not txn:
        flash('Transaction not found', 'error')
        return redirect(url_for('home'))
    accounts = models.get_accounts()
    company = models.get_meta('company_name', 'My Books')
    return_to = request.args.get('return_to', '')
    from_account_id = request.args.get('from_account', '')
    from_report_id = request.args.get('from_report', '')
    
    # Build breadcrumb context
    from_account = models.get_account(int(from_account_id)) if from_account_id else None
    from_report = models.get_report(int(from_report_id)) if from_report_id else None
    
    return render_template('edit_transaction.html', txn=txn, lines=lines,
                         accounts=accounts, company=company, return_to=return_to,
                         from_account=from_account, from_report=from_report)

# ─── Delete Transaction ────────────────────────────────────────────

@app.route('/transaction/<int:txn_id>/delete', methods=['POST'])
def delete_transaction(txn_id):
    try:
        models.delete_transaction(txn_id)
        flash('Transaction deleted', 'success')
    except ValueError as e:
        flash(str(e), 'error')
    return_to = request.form.get('return_to', url_for('home'))
    return redirect(return_to)

@app.route('/api/bulk-delete', methods=['POST'])
def api_bulk_delete():
    """Delete multiple transactions at once."""
    try:
        data = request.get_json()
        txn_ids = data.get('txn_ids', [])
        if not txn_ids:
            return jsonify({'ok': False, 'error': 'No transactions selected'})
        deleted, skipped = models.bulk_delete_transactions(txn_ids)
        msg = f'Deleted {deleted} transaction{"s" if deleted != 1 else ""}'
        if skipped:
            msg += f' ({skipped} skipped — before lock date)'
        return jsonify({'ok': True, 'deleted': deleted, 'skipped': skipped, 'message': msg})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/gst-split/<int:txn_id>', methods=['POST'])
def api_gst_split(txn_id):
    """Split an existing 2-line transaction into 3 lines with GST.
    F8 = purchase (GST.IN / ITC), F9 = sale (GST.OUT / collected).
    Rate and accounts are configurable in Options."""
    try:
        data = request.get_json()
        split_type = data.get('type', 'purchase')  # 'purchase' or 'sale'
        
        txn, lines = models.get_transaction(txn_id)
        if not txn:
            return jsonify({'ok': False, 'error': 'Transaction not found'})
        
        if len(lines) != 2:
            return jsonify({'ok': False, 'error': f'Can only split simple 2-line transactions (this has {len(lines)} lines)'})
        
        lock = models.get_meta('lock_date', '')
        if lock and txn['date'] <= lock:
            return jsonify({'ok': False, 'error': f'Transaction is locked (before {lock})'})
        
        # Load configurable rate and accounts
        rate_num = int(models.get_meta('gst_rate_num', '5'))
        rate_den = int(models.get_meta('gst_rate_den', '105'))
        
        if split_type == 'purchase':
            gst_acct_name = models.get_meta('f8_tax_acct', 'GST.IN')
            post_acct_name = models.get_meta('f8_post_acct', '')
        else:
            gst_acct_name = models.get_meta('f9_tax_acct', 'GST.OUT')
            post_acct_name = models.get_meta('f9_post_acct', '')
        
        # Figure out which line is the bank (the account we're viewing) and which is the cross
        from_account = int(data.get('from_account', 0))
        bank_line = None
        cross_line = None
        for ln in lines:
            if ln['account_id'] == from_account:
                bank_line = ln
            else:
                cross_line = ln
        
        if not bank_line or not cross_line:
            bank_line = lines[0]
            cross_line = lines[1]
        
        # Calculate tax from gross: tax = gross * rate_num / rate_den
        gross = abs(cross_line['amount'])  # in cents
        tax = round(gross * rate_num / rate_den)
        net = gross - tax
        
        # Resolve accounts
        gst_acct = models.get_account_by_name(gst_acct_name)
        if not gst_acct:
            return jsonify({'ok': False, 'error': f'Tax account "{gst_acct_name}" not found. Set it up in Options.'})
        
        # If a default posting account is configured, re-assign the cross line
        cross_acct_id = cross_line['account_id']
        if post_acct_name:
            post_acct = models.get_account_by_name(post_acct_name)
            if post_acct:
                cross_acct_id = post_acct['id']
        
        # Preserve the sign of the cross line
        sign = 1 if cross_line['amount'] > 0 else -1
        
        new_lines = [
            (bank_line['account_id'], bank_line['amount'], bank_line['description']),
            (cross_acct_id, sign * net, cross_line['description']),
            (gst_acct['id'], sign * tax, f'GST: {cross_line["description"][:40]}'),
        ]
        
        # Verify balance
        total = sum(l[1] for l in new_lines)
        if total != 0:
            return jsonify({'ok': False, 'error': f'Split does not balance (off by {total/100:.2f})'})
        
        models.update_transaction(txn_id, txn['date'], txn['reference'], txn['description'], new_lines)
        
        return jsonify({'ok': True, 'txn_id': txn_id})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# ─── Reconcile Toggle ──────────────────────────────────────────────

@app.route('/transaction/<int:txn_id>/reconcile', methods=['POST'])
def reconcile(txn_id):
    line_id = request.form.get('line_id', 0, type=int)
    if line_id:
        models.toggle_reconcile(line_id)
    return_to = request.form.get('return_to', url_for('home'))
    return redirect(return_to)

# ─── Reconciliation View ─────────────────────────────────────────

@app.route('/reconcile/<int:account_id>')
def reconcile_view(account_id):
    """Bank reconciliation screen."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    account = models.get_account(account_id)
    if not account:
        flash('Account not found', 'error')
        return redirect(url_for('home'))
    
    stmt_bal_str = request.args.get('stmt', '')
    stmt_bal = models.parse_amount(stmt_bal_str) if stmt_bal_str.strip() else None
    
    entries = models.get_ledger(account_id)
    sign = 1 if account['normal_balance'] == 'D' else -1
    summary = models.get_reconcile_summary(account_id)
    company = models.get_meta('company_name', 'My Books')
    
    # Separate cleared vs outstanding
    outstanding = [e for e in entries if not e['reconciled'] and e['amount'] != 0]
    cleared = [e for e in entries if e['reconciled'] and e['amount'] != 0]
    
    # Calculate difference
    diff = None
    if stmt_bal is not None:
        # Statement balance should equal cleared balance
        # For a bank (debit normal): stmt balance = cleared balance
        diff = (stmt_bal * sign) - summary['cleared_balance']
    
    from_report = request.args.get('from_report', '')
    parent_report = models.get_report(int(from_report)) if from_report else None
    
    return render_template('reconcile.html', account=account,
        outstanding=outstanding, cleared=cleared, summary=summary,
        stmt_bal=stmt_bal_str, diff=diff, company=company,
        parent_report=parent_report, from_report=from_report)

@app.route('/api/reconcile-toggle/<int:line_id>', methods=['POST'])
def api_reconcile_toggle(line_id):
    """AJAX toggle for reconciliation checkboxes."""
    new_val = models.toggle_reconcile(line_id)
    # Return updated summary for the account
    with models.get_db() as db:
        row = db.execute("SELECT account_id FROM lines WHERE id=?", (line_id,)).fetchone()
    summary = models.get_reconcile_summary(row['account_id'])
    return jsonify({'ok': True, 'reconciled': new_val,
                   'cleared_balance': summary['cleared_balance'],
                   'uncleared': summary['uncleared'],
                   'book_balance': summary['book_balance']})

@app.route('/api/doc-toggle/<int:line_id>', methods=['POST'])
def api_doc_toggle(line_id):
    """AJAX toggle for document-on-file flag."""
    new_val = models.toggle_doc_on_file(line_id)
    return jsonify({'ok': True, 'doc_on_file': new_val})

# ─── Trial Balance ──────────────────────────────────────────────────

@app.route('/trial-balance')
def trial_balance():
    if not models.get_db_path():
        return redirect(url_for('library'))
    as_of = request.args.get('as_of', '')
    accounts, total_dr, total_cr = models.get_trial_balance(as_of or None)
    company = models.get_meta('company_name', 'My Books')
    return render_template('trial_balance.html', accounts=accounts,
                         total_dr=total_dr, total_cr=total_cr,
                         as_of=as_of, company=company)

# ─── Search ─────────────────────────────────────────────────────────

@app.route('/search')
def search():
    if not models.get_db_path():
        return redirect(url_for('library'))
    query = request.args.get('q', '')
    results = models.search_transactions(query) if query else []
    company = models.get_meta('company_name', 'My Books')
    return render_template('search.html', query=query, results=results,
                         company=company)

# ─── Account Management ────────────────────────────────────────────

@app.route('/account/new', methods=['GET', 'POST'])
@app.route('/account/new/in/<int:report_id>', methods=['GET', 'POST'])
def new_account(report_id=None):
    if not models.get_db_path():
        return redirect(url_for('library'))
    if request.method == 'POST':
        name = request.form['name'].strip().upper()
        desc = request.form.get('description', '')
        nb = request.form.get('normal_balance', 'D')
        acct_num = request.form.get('account_number', '')
        try:
            acct_id = models.add_account(name, nb, desc, account_number=acct_num)
            flash(f'Account {name} created', 'success')
            return redirect(url_for('home'))
        except Exception as e:
            flash(str(e), 'error')
    reports = models.get_reports()
    company = models.get_meta('company_name', 'My Books')
    return render_template('new_account.html', reports=reports, 
                         report_id=report_id, company=company)

# ─── Report Management ─────────────────────────────────────────────

@app.route('/report/new', methods=['GET', 'POST'])
def new_report():
    if not models.get_db_path():
        return redirect(url_for('library'))
    if request.method == 'POST':
        name = request.form['name'].strip()
        desc = request.form.get('description', '')
        try:
            rid = models.add_report(name, desc)
            # Auto-seed with blank lines so the report is never empty
            models.add_report_item(rid, 'label', desc.upper() if desc else name.upper(), position=10, indent=0)
            models.add_report_item(rid, 'label', '', position=20)
            models.add_report_item(rid, 'separator', position=30, sep_style='double')
            flash(f'Report "{name}" created', 'success')
            return redirect(url_for('report_view', report_id=rid))
        except Exception as e:
            flash(str(e), 'error')
    company = models.get_meta('company_name', 'My Books')
    return render_template('new_report.html', company=company)

# ─── Settings ───────────────────────────────────────────────────────

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if not models.get_db_path():
        return redirect(url_for('library'))
    if request.method == 'POST':
        models.set_meta('company_name', request.form.get('company_name', ''))
        models.set_meta('fiscal_year_end', request.form.get('fiscal_year_end', ''))
        models.set_meta('fiscal_year', request.form.get('fiscal_year', ''))
        models.set_meta('lock_date', request.form.get('lock_date', ''))
        # F8/F9 GST split settings
        models.set_meta('gst_rate_num', request.form.get('gst_rate_num', '5'))
        models.set_meta('gst_rate_den', request.form.get('gst_rate_den', '105'))
        models.set_meta('f8_tax_acct', request.form.get('f8_tax_acct', 'GST.IN'))
        models.set_meta('f8_post_acct', request.form.get('f8_post_acct', ''))
        models.set_meta('f9_tax_acct', request.form.get('f9_tax_acct', 'GST.OUT'))
        models.set_meta('f9_post_acct', request.form.get('f9_post_acct', ''))
        # Stripe procedure settings
        models.set_meta('stripe_fee_acct', request.form.get('stripe_fee_acct', 'EX.CC'))
        models.set_meta('stripe_ar_report', request.form.get('stripe_ar_report', ''))
        flash('Settings saved', 'success')
        return redirect(url_for('settings'))
    company = models.get_meta('company_name', 'My Books')
    fye = models.get_meta('fiscal_year_end', '12-31')
    fy_year = models.get_meta('fiscal_year', '')
    lock_date = models.get_meta('lock_date', '')
    db_path = models.get_db_path()
    # F8/F9 settings
    gst_rate_num = models.get_meta('gst_rate_num', '5')
    gst_rate_den = models.get_meta('gst_rate_den', '105')
    f8_tax_acct = models.get_meta('f8_tax_acct', 'GST.IN')
    f8_post_acct = models.get_meta('f8_post_acct', '')
    f9_tax_acct = models.get_meta('f9_tax_acct', 'GST.OUT')
    f9_post_acct = models.get_meta('f9_post_acct', '')
    # Stripe settings
    stripe_fee_acct = models.get_meta('stripe_fee_acct', 'EX.CC')
    stripe_ar_report = models.get_meta('stripe_ar_report', '')
    accounts = models.get_accounts()
    reports = models.get_reports()
    return render_template('settings.html', company=company, fye=fye, fy_year=fy_year,
                         lock_date=lock_date, db_path=db_path,
                         gst_rate_num=gst_rate_num, gst_rate_den=gst_rate_den,
                         f8_tax_acct=f8_tax_acct, f8_post_acct=f8_post_acct,
                         f9_tax_acct=f9_tax_acct, f9_post_acct=f9_post_acct,
                         stripe_fee_acct=stripe_fee_acct, stripe_ar_report=stripe_ar_report,
                         accounts=accounts, reports=reports)

# ─── CSV Import ──────────────────────────────────────────────────

def _fix_csv_rows(all_rows):
    """Fix rows where unquoted commas in fields cause extra columns.
    Strategy: header defines N columns. For rows with >N cols, we know the
    rightmost columns are numeric (amounts) and the excess splits happened
    in text fields. We anchor from the right (amounts) and merge the middle."""
    if not all_rows:
        return all_rows
    ncols = len(all_rows[0])
    fixed = [all_rows[0]]
    for row in all_rows[1:]:
        if len(row) <= ncols:
            fixed.append(row)
            continue
        excess = len(row) - ncols
        # Take the first few columns as-is (account type, number, date, ref)
        # These are never the ones that split. Then merge the description fields.
        # Take the last columns as-is (amounts, empty trailing fields).
        # Strategy: keep first 4 cols, keep last (ncols-5) cols, merge everything in between.
        # For an 8-col header: keep [0:4], merge [4:4+1+excess], keep [-3:]
        # This merges the Description fields (cols 4 and 5 in the original)
        left_keep = 4  # Account Type, Account Number, Date, Cheque#
        right_keep = ncols - left_keep - 2  # Amount cols + trailing (usually 2-3 cols)
        # But we need at least the desc cols (normally 2) to end up as ncols - left_keep - right_keep
        middle_start = left_keep
        middle_end = len(row) - right_keep
        middle = row[middle_start:middle_end]
        # middle should be exactly 2 fields (desc1, desc2) but has 2+excess
        # Merge all middle fields into exactly 2: first field stays, rest merge
        if len(middle) >= 2:
            desc1 = middle[0]
            desc2 = ', '.join(middle[1:])
            new_row = row[:left_keep] + [desc1, desc2] + row[middle_end:]
        else:
            new_row = row[:ncols]
        fixed.append(new_row[:ncols])
    return fixed

@app.route('/import', methods=['GET', 'POST'])
def csv_import():
    if not models.get_db_path():
        return redirect(url_for('library'))
    company = models.get_meta('company_name', 'My Books')
    accounts = models.get_accounts()
    import json as _json, tempfile, hashlib
    
    # Cache helpers — store import data in temp files to avoid cookie size limits
    cache_dir = os.path.join(tempfile.gettempdir(), 'grid_import')
    os.makedirs(cache_dir, exist_ok=True)
    
    def _cache_key():
        """Get or create a per-session cache key."""
        k = session.get('import_cache_key')
        if not k:
            k = hashlib.md5(os.urandom(16)).hexdigest()[:12]
            session['import_cache_key'] = k
        return k
    
    def _cache_put(name, data):
        path = os.path.join(cache_dir, f'{_cache_key()}_{name}.json')
        with open(path, 'w') as f:
            _json.dump(data, f)
    
    def _cache_get(name):
        path = os.path.join(cache_dir, f'{_cache_key()}_{name}.json')
        if os.path.exists(path):
            with open(path) as f:
                return _json.load(f)
        return None
    
    def _cache_clear():
        k = session.get('import_cache_key', '')
        if k:
            for fn in os.listdir(cache_dir):
                if fn.startswith(k):
                    os.remove(os.path.join(cache_dir, fn))
            session.pop('import_cache_key', None)
    
    step = request.form.get('_step', '')
    
    # ═══ STEP 3: CONFIRM IMPORT ═══
    if step == 'do_import':
        cached = _cache_get('parsed')
        if not cached:
            flash('Session expired — please re-upload the file', 'error')
            return redirect(url_for('csv_import'))
        
        account_id = int(session.get('import_account_id', 0))
        acct = models.get_account(account_id)
        if not acct:
            flash('Account not found', 'error')
            return redirect(url_for('csv_import'))
        
        susp = models.get_account_by_name('EX.SUSP')
        susp_id = susp['id'] if susp else models.add_account('EX.SUSP', 'D', 'Suspense')
        nb_flip = -1 if acct['normal_balance'] == 'C' else 1
        
        count = 0
        try:
            for item in cached:
                d, ref, desc, amount = item['date'], item['ref'], item['desc'], item['amount']
                post_amount = amount * nb_flip
                rule_acct, rule_tax, tax_split = models.apply_rules(desc, amount)
                cross = models.get_account_by_name(rule_acct)
                cross_id = cross['id'] if cross else susp_id
                
                if tax_split:
                    tax_acct = models.get_account_by_name(tax_split['tax_acct'])
                    tax_acct_id = tax_acct['id'] if tax_acct else susp_id
                    net, tax = tax_split['net'], tax_split['tax']
                    if post_amount > 0:
                        models.add_transaction(d, ref, desc, [
                            (account_id, post_amount, desc),
                            (cross_id, -net, desc),
                            (tax_acct_id, -tax, f'Tax: {desc[:40]}')])
                    else:
                        models.add_transaction(d, ref, desc, [
                            (account_id, post_amount, desc),
                            (cross_id, net, desc),
                            (tax_acct_id, tax, f'Tax: {desc[:40]}')])
                else:
                    if post_amount > 0:
                        models.add_simple_transaction(d, ref, desc, account_id, cross_id, post_amount)
                    else:
                        models.add_simple_transaction(d, ref, desc, cross_id, account_id, abs(post_amount))
                count += 1
            
            _cache_clear()
            flash(f'Imported {count} transactions into {acct["name"]}', 'success')
            return redirect(url_for('account_ledger', account_id=account_id))
        except Exception as e:
            flash(f'Import failed: {e}', 'error')
            return redirect(url_for('csv_import'))
    
    # ═══ STEP 2b: REFRESH PREVIEW (re-apply rules after adding new ones) ═══
    if step == 'refresh_preview':
        cached = _cache_get('parsed')
        if not cached:
            flash('Session expired — please re-upload the file', 'error')
            return redirect(url_for('csv_import'))
        
        account_id = int(session.get('import_account_id', 0))
        acct = models.get_account(account_id)
        if not acct:
            flash('Account not found', 'error')
            return redirect(url_for('csv_import'))
        
        # Re-apply rules to the already-parsed data
        preview = []
        for item in cached:
            rule_acct, rule_tax, tax_split = models.apply_rules(item['desc'], item['amount'])
            preview.append({**item,
                'rule_acct': rule_acct, 'rule_tax': rule_tax, 'tax_split': tax_split})
        
        matched = sum(1 for p in preview if p['rule_acct'] != 'EX.SUSP')
        return render_template('import.html', company=company, accounts=accounts,
            rules_preview=preview, account_id=account_id,
            matched_count=matched, acct_name=acct['name'], tax_codes=models.get_tax_codes())
    
    # ═══ STEP 2: PARSE WITH DATE FORMAT + SHOW RULES PREVIEW ═══
    if step == 'preview':
        cached_rows = _cache_get('rows')
        if not cached_rows:
            flash('Session expired — please re-upload the file', 'error')
            return redirect(url_for('csv_import'))
        
        account_id = int(request.form.get('account_id', session.get('import_account_id', 0)))
        session['import_account_id'] = account_id
        acct = models.get_account(account_id)
        if not acct:
            flash('Account not found', 'error')
            return redirect(url_for('csv_import'))
        
        date_fmt = request.form.get('date_format', 'MDY')
        col_date = int(request.form.get('col_date', 0))
        col_ref = int(request.form.get('col_ref', -1))
        col_desc = int(request.form.get('col_desc', 0))
        col_desc2 = int(request.form.get('col_desc2', -1))
        col_amount = int(request.form.get('col_amount', -1))
        col_debit = int(request.form.get('col_debit', -1))
        col_credit = int(request.form.get('col_credit', -1))
        skip_header = request.form.get('skip_header', '0') == '1'
        flip_sign = request.form.get('flip_sign', '0') == '1'
        
        data_rows = cached_rows[1:] if skip_header else cached_rows
        lock = models.get_meta('lock_date', '')
        errors = []
        parsed = []
        
        for i, row in enumerate(data_rows):
            line_num = i + (2 if skip_header else 1)
            try:
                if not row or all(str(c).strip() == '' for c in row):
                    continue
                def col(idx, _row=row):
                    if idx < 0 or idx >= len(_row): return ''
                    return str(_row[idx]).strip()
                
                d = _parse_date(col(col_date), fmt=date_fmt)
                if not d:
                    errors.append(f"Row {line_num}: Invalid date '{col(col_date)}'")
                    continue
                
                # Validate year range
                try:
                    yr = int(d[:4])
                    if yr < 1950 or yr > 2099:
                        errors.append(f"Row {line_num}: Year {yr} out of range (1950-2099)")
                        continue
                except:
                    pass
                
                if lock and d <= lock:
                    errors.append(f"Row {line_num}: Date {d} is on or before lock date ({lock})")
                    continue
                
                ref = col(col_ref)
                desc = col(col_desc)
                if col_desc2 >= 0:
                    d2 = col(col_desc2)
                    if d2: desc = f"{desc} — {d2}" if desc else d2
                
                if col_amount >= 0:
                    amt_str = col(col_amount).replace(',', '').replace('$', '').replace('"', '')
                    if not amt_str: continue
                    amount = models.parse_amount(amt_str)
                else:
                    dr_str = col(col_debit).replace(',', '').replace('$', '').replace('"', '')
                    cr_str = col(col_credit).replace(',', '').replace('$', '').replace('"', '')
                    dr_val = models.parse_amount(dr_str) if dr_str else 0
                    cr_val = models.parse_amount(cr_str) if cr_str else 0
                    amount = dr_val - cr_val
                
                if amount == 0: continue
                if flip_sign: amount = -amount
                parsed.append({'date': d, 'ref': ref, 'desc': desc, 'amount': amount})
            except Exception as e:
                errors.append(f"Row {line_num}: {e}")
        
        if errors and not parsed:
            flash(f"{len(errors)} error(s) found, no valid transactions.", 'error')
            return render_template('import.html', company=company, accounts=accounts,
                errors=errors)
        
        if not parsed:
            flash('No valid transactions found in file', 'error')
            return redirect(url_for('csv_import'))
        
        # Cache parsed data for the confirm step
        _cache_put('parsed', parsed)
        
        # Build rules preview
        preview = []
        for item in parsed:
            rule_acct, rule_tax, tax_split = models.apply_rules(item['desc'], item['amount'])
            preview.append({**item,
                'rule_acct': rule_acct, 'rule_tax': rule_tax, 'tax_split': tax_split})
        
        matched = sum(1 for p in preview if p['rule_acct'] != 'EX.SUSP')
        return render_template('import.html', company=company, accounts=accounts,
            rules_preview=preview, account_id=account_id,
            matched_count=matched, errors=errors if errors else None,
            acct_name=acct['name'], tax_codes=models.get_tax_codes())
    
    # ═══ STEP 1: UPLOAD FILE → SHOW COLUMN MAPPING + DATE DETECTION ═══
    if request.method == 'POST':
        account_id = int(request.form.get('account_id', 0))
        session['import_account_id'] = account_id
        
        file = request.files.get('csv_file')
        if not file or not file.filename:
            flash('No file selected', 'error')
            return redirect(url_for('csv_import'))
        
        fname = file.filename.lower()
        if not any(fname.endswith(ext) for ext in ('.csv', '.xlsx', '.xls', '.txt', '.tsv')):
            flash('Unsupported file type. Use .csv, .xlsx, or .xls', 'error')
            return redirect(url_for('csv_import'))
        
        try:
            all_rows = _read_upload_to_rows(file)
            all_rows = _fix_csv_rows(all_rows)
        except Exception as e:
            flash(f'Error reading file: {e}', 'error')
            return redirect(url_for('csv_import'))
        
        if len(all_rows) < 2:
            flash('File has no data rows', 'error')
            return redirect(url_for('csv_import'))
        
        # Cache rows in temp file
        _cache_put('rows', all_rows)
        
        # Auto-detect date column and format
        headers = all_rows[0]
        sample_rows = all_rows[1:min(51, len(all_rows))]
        
        best_date_col = 0
        best_date_score = 0
        for ci in range(len(headers)):
            samples = [str(row[ci]).strip() for row in sample_rows if ci < len(row) and str(row[ci]).strip()]
            score = sum(1 for s in samples if _parse_date(s) is not None)
            if score > best_date_score:
                best_date_score = score
                best_date_col = ci
        
        date_samples = [str(row[best_date_col]).strip() for row in sample_rows
                       if best_date_col < len(row) and str(row[best_date_col]).strip()]
        date_fmt, confidence, detail = _detect_date_format(date_samples)
        
        # Parse samples under each format for side-by-side display
        date_comparisons = []
        has_ambiguous = False  # only show MDY/DMY columns if there are ambiguous numeric dates
        for raw in date_samples[:10]:
            parsed_auto = _parse_date(raw, 'auto')
            parsed_mdy = _parse_date(raw, 'MDY')
            parsed_dmy = _parse_date(raw, 'DMY')
            # Check if MDY and DMY give different results (ambiguous)
            if parsed_mdy and parsed_dmy and parsed_mdy != parsed_dmy:
                has_ambiguous = True
            # "Parsed as" uses the detected format
            parsed = _parse_date(raw, date_fmt)
            date_comparisons.append({
                'raw': raw,
                'parsed': parsed,
                'MDY': parsed_mdy,
                'DMY': parsed_dmy,
                'ok': parsed is not None,
            })
        
        parse_ok = sum(1 for dc in date_comparisons if dc['ok'])
        parse_fail = len(date_comparisons) - parse_ok
        # Also count across ALL dates (not just samples)
        all_parse_ok = sum(1 for raw in date_samples if _parse_date(raw, date_fmt) is not None)
        all_parse_fail = len(date_samples) - all_parse_ok
        
        preview_rows = all_rows[1:6] if len(all_rows) > 1 else []
        
        return render_template('import.html', company=company, accounts=accounts,
            headers=headers, preview_rows=preview_rows, num_cols=len(headers),
            account_id=account_id, errors=None,
            date_fmt=date_fmt, date_confidence=confidence, date_detail=detail,
            date_comparisons=date_comparisons, best_date_col=best_date_col,
            total_rows=len(all_rows) - 1,
            has_ambiguous=has_ambiguous,
            all_parse_ok=all_parse_ok, all_parse_fail=all_parse_fail)
    
    # ═══ INITIAL: Upload form ═══
    _cache_clear()
    return render_template('import.html', company=company, accounts=accounts)

def _parse_date(s, fmt='auto'):
    """Parse a date string into yyyy-mm-dd format.
    fmt: 'auto' (guess), 'MDY', 'DMY', 'YMD'.
    Handles: ISO, slashes, dashes, dots, month names (short/long), spaces, 2/4-digit years."""
    import re
    s = str(s).strip().strip('"').strip("'")
    if not s or s.lower() in ('', 'none', 'nat', 'null'):
        return None
    
    months = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
              'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12,
              'january':1,'february':2,'march':3,'april':4,'june':6,
              'july':7,'august':8,'september':9,'october':10,'november':11,'december':12}
    
    def _fix_year(y):
        if y < 100:
            return y + 2000 if y < 50 else y + 1900
        return y
    
    def _valid(y, m, d):
        """Check if date components are valid."""
        if m < 1 or m > 12 or d < 1 or d > 31:
            return None
        try:
            from datetime import datetime
            datetime(y, m, d)
            return f"{y:04d}-{m:02d}-{d:02d}"
        except ValueError:
            return None
    
    # ── Unambiguous formats (always try first regardless of fmt) ──
    
    # ISO: yyyy-mm-dd or yyyy/mm/dd or yyyy.mm.dd
    m = re.match(r'^(\d{4})[\-/.](\d{1,2})[\-/.](\d{1,2})$', s)
    if m:
        return _valid(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    
    # yyyymmdd (compact ISO)
    m = re.match(r'^(\d{4})(\d{2})(\d{2})$', s)
    if m:
        return _valid(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    
    # ── Month name formats (unambiguous — month is spelled out) ──
    
    # dd Mon yyyy or dd-Mon-yyyy or dd/Mon/yyyy or dd.Mon.yyyy (e.g. "31 May 2025", "15-Jan-2025")
    m = re.match(r'^(\d{1,2})[\s\-/.](\w+)[\s\-/.,]+(\d{2,4})$', s, re.I)
    if m and m.group(2).lower() in months:
        return _valid(_fix_year(int(m.group(3))), months[m.group(2).lower()], int(m.group(1)))
    
    # Mon dd, yyyy or Mon dd yyyy (e.g. "May 31, 2025", "Jan 15 2025")
    m = re.match(r'^(\w+)[\s\-/.](\d{1,2})[,\s]+(\d{2,4})$', s, re.I)
    if m and m.group(1).lower() in months:
        return _valid(_fix_year(int(m.group(3))), months[m.group(1).lower()], int(m.group(2)))
    
    # yyyy Mon dd or yyyy-Mon-dd (e.g. "2025 May 31")
    m = re.match(r'^(\d{4})[\s\-/.](\w+)[\s\-/.](\d{1,2})$', s, re.I)
    if m and m.group(2).lower() in months:
        return _valid(int(m.group(1)), months[m.group(2).lower()], int(m.group(3)))
    
    # ── Numeric formats (ambiguous — depends on fmt) ──
    
    # nn/nn/nnnn or nn-nn-nnnn or nn.nn.nnnn
    m = re.match(r'^(\d{1,2})[\-/.](\d{1,2})[\-/.](\d{4})$', s)
    if m:
        a, b, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if fmt == 'DMY':
            return _valid(yr, b, a)
        else:  # MDY or auto
            return _valid(yr, a, b)
    
    # nn/nn/nn (2-digit year)
    m = re.match(r'^(\d{1,2})[\-/.](\d{1,2})[\-/.](\d{2})$', s)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        yr = _fix_year(int(m.group(3)))
        if fmt == 'DMY':
            return _valid(yr, b, a)
        else:
            return _valid(yr, a, b)
    
    return None

def _detect_date_format(date_strings):
    """Auto-detect date format from sample strings. Returns (fmt, confidence, detail).
    fmt is only meaningful for ambiguous numeric dates (MDY vs DMY).
    Month-name and ISO dates are always unambiguous."""
    import re
    if not date_strings:
        return 'MDY', 'low', 'no dates found'
    
    # Count how many parse successfully with auto
    auto_ok = sum(1 for d in date_strings if _parse_date(d, 'auto') is not None)
    
    # YMD: year comes first
    ymd_count = sum(1 for d in date_strings if re.match(r'^\d{4}[/\-.]', str(d)))
    if ymd_count > len(date_strings) * 0.8:
        return 'YMD', 'high', f'{ymd_count}/{len(date_strings)} start with 4-digit year'
    
    # Month names are unambiguous
    month_count = sum(1 for d in date_strings if re.search(r'[A-Za-z]{3,}', str(d)))
    if month_count > len(date_strings) * 0.8:
        return 'auto', 'high', f'month names detected — dates are unambiguous ({auto_ok}/{len(date_strings)} parsed OK)'
    
    # MDY vs DMY: check if first or second component > 12
    firsts, seconds = [], []
    for d in date_strings:
        m = re.match(r'^(\d{1,2})[/\-.](\d{1,2})[/\-.]', str(d))
        if m:
            firsts.append(int(m.group(1)))
            seconds.append(int(m.group(2)))
    
    if not firsts:
        return 'MDY', 'low', 'could not parse date components'
    
    first_over_12 = any(v > 12 for v in firsts)
    second_over_12 = any(v > 12 for v in seconds)
    
    if first_over_12 and not second_over_12:
        return 'DMY', 'high', f'first component goes up to {max(firsts)} (must be day)'
    if second_over_12 and not first_over_12:
        return 'MDY', 'high', f'second component goes up to {max(seconds)} (must be day)'
    
    return 'MDY', 'ambiguous', 'all values ≤ 12 — could be MDY or DMY. Defaulting to MDY (North American).'

def _read_upload_to_rows(fileobj):
    """Read uploaded file (csv, xlsx, xls) into list of lists of strings."""
    import csv, io
    fname = fileobj.filename.lower()
    
    if fname.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(fileobj, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(v) if v is not None else '' for v in row])
        return rows
    elif fname.endswith('.xls'):
        # Old Excel 97-2003 binary format — needs xlrd
        try:
            import xlrd
        except ImportError:
            raise ImportError('xlrd is required for .xls files. Run: pip install xlrd')
        raw = fileobj.read()
        wb = xlrd.open_workbook(file_contents=raw)
        ws = wb.sheet_by_index(0)
        rows = []
        for rx in range(ws.nrows):
            row = []
            for cx in range(ws.ncols):
                cell = ws.cell(rx, cx)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    # Convert Excel date to string
                    try:
                        dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        row.append(dt.strftime('%Y-%m-%d'))
                    except:
                        row.append(str(cell.value))
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    # Keep numbers clean (no trailing .0 for integers)
                    v = cell.value
                    row.append(str(int(v)) if v == int(v) else str(v))
                else:
                    row.append(str(cell.value) if cell.value is not None else '')
            rows.append(row)
        return rows
    else:
        # CSV — handle various encodings
        raw = fileobj.read()
        for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                content = raw.decode(enc)
                break
            except (UnicodeDecodeError, AttributeError):
                continue
        else:
            content = raw.decode('utf-8', errors='replace')
        
        reader = csv.reader(io.StringIO(content))
        return [row for row in reader]

# ─── Quick Entry API (inline ledger posting) ───────────────────────

@app.route('/api/quick-entry/<int:account_id>', methods=['POST'])
def api_quick_entry(account_id):
    """Post a transaction from the inline ledger entry row."""
    try:
        date_str = request.form.get('date', '')
        reference = request.form.get('reference', '')
        description = request.form.get('description', '')
        amount_str = request.form.get('amount', '0')
        cross_name = request.form.get('cross_account', '').strip().upper()
        gst_split = request.form.get('gst_split', '0') == '1'
        
        if not date_str:
            return jsonify({'ok': False, 'error': 'Date required'})
        
        # Validate and normalize date
        parsed_date = _parse_date(date_str)
        if not parsed_date:
            return jsonify({'ok': False, 'error': 'Invalid date format. Use yyyy-mm-dd (e.g. 2025-01-15)'})
        date_str = parsed_date
        
        amount = models.parse_amount(amount_str) if amount_str.strip() else 0
        this_acct = models.get_account(account_id)
        
        # Flip sign for credit-normal accounts: user types positive to mean
        # "increase this account" which for a credit-normal account means a credit (negative internally)
        if this_acct['normal_balance'] == 'C' and amount != 0:
            amount = -amount
        
        # Blank/placeholder entry: date only, no amount, no cross-account
        # Creates a single zero-amount line (memo/separator)
        if amount == 0 and not cross_name:
            if not reference or not reference.strip():
                reference = models.generate_ref()
            with models.get_db() as db:
                cur = db.execute("INSERT INTO transactions(date, reference, description) VALUES(?,?,?)",
                    (date_str, reference, description))
                txn_id = cur.lastrowid
                db.execute("INSERT INTO lines(transaction_id, account_id, amount, description, sort_order) VALUES(?,?,0,?,0)",
                    (txn_id, account_id, description))
            return jsonify({'ok': True})
        
        if gst_split and amount != 0:
            gst_type = request.form.get('gst_type', 'purchase')
            
            # Load configurable rate and accounts from settings
            rate_num = int(models.get_meta('gst_rate_num', '5'))
            rate_den = int(models.get_meta('gst_rate_den', '105'))
            
            if gst_type == 'sale':
                gst_acct_name = models.get_meta('f9_tax_acct', 'GST.OUT')
                default_post = models.get_meta('f9_post_acct', '')
            else:
                gst_acct_name = models.get_meta('f8_tax_acct', 'GST.IN')
                default_post = models.get_meta('f8_post_acct', '')
            
            # Use the cross-account if provided, otherwise the configured default, otherwise EX.SUSP
            if cross_name:
                cross_acct = models.get_account_by_name(cross_name)
                if not cross_acct:
                    return jsonify({'ok': False, 'error': f'Account not found: {cross_name}'})
            elif default_post:
                cross_acct = models.get_account_by_name(default_post)
                if not cross_acct:
                    return jsonify({'ok': False, 'error': f'Default posting account "{default_post}" not found. Check Options.'})
            else:
                cross_acct = models.get_account_by_name('EX.SUSP')
                if not cross_acct:
                    return jsonify({'ok': False, 'error': 'EX.SUSP account not found'})
            
            gst_acct = models.get_account_by_name(gst_acct_name)
            if not gst_acct:
                return jsonify({'ok': False, 'error': f'Tax account "{gst_acct_name}" not found. Check Options.'})
            
            abs_amt = abs(amount)
            gst_cents = round(abs_amt * rate_num / rate_den)
            net_cents = abs_amt - gst_cents
            gst_desc = 'GST collected' if gst_type == 'sale' else 'ITCs paid'
            
            if amount < 0:
                lines = [
                    (account_id, amount, description),
                    (cross_acct['id'], net_cents, description),
                    (gst_acct['id'], gst_cents, gst_desc),
                ]
            else:
                lines = [
                    (account_id, amount, description),
                    (cross_acct['id'], -net_cents, description),
                    (gst_acct['id'], -gst_cents, gst_desc),
                ]
            
            models.add_transaction(date_str, reference, description, lines)
            return jsonify({'ok': True})
        
        if cross_name:
            cross_acct = models.get_account_by_name(cross_name)
            if not cross_acct:
                return jsonify({'ok': False, 'error': f'Account not found: {cross_name}'})
            
            if amount > 0:
                models.add_simple_transaction(date_str, reference, description,
                    account_id, cross_acct['id'], amount)
            elif amount < 0:
                models.add_simple_transaction(date_str, reference, description,
                    cross_acct['id'], account_id, abs(amount))
            else:
                return jsonify({'ok': False, 'error': 'Amount cannot be zero'})
        else:
            return jsonify({'ok': False, 'error': 'Cross-account required'})
        
        return jsonify({'ok': True})
        
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/account-search')
def api_account_search():
    q = request.args.get('q', '')
    posting_only = request.args.get('posting', '0') == '1'
    if len(q) < 1:
        return jsonify([])
    accounts = models.search_accounts(q)
    results = []
    for a in accounts:
        if posting_only and a['account_type'] == 'total':
            continue
        results.append({'name': a['name'], 'description': a['description'], 'id': a['id']})
    return jsonify(results)

@app.route('/api/inline-edit', methods=['POST'])
def api_inline_edit():
    """Edit a single field on an existing transaction/line — inline editing."""
    try:
        data = request.get_json()
        txn_id = data.get('txn_id')
        line_id = data.get('line_id')
        field = data.get('field')
        value = data.get('value', '')
        
        with models.get_db() as db:
            lock = models.get_meta('lock_date', '')
            txn = db.execute("SELECT date FROM transactions WHERE id=?", (txn_id,)).fetchone()
            
            # For non-date fields: reject if transaction is already locked
            if field != 'date' and lock and txn and txn['date'] <= lock:
                return jsonify({'ok': False, 'error': f'Transaction is locked (before {lock})'})
            
            if field == 'date':
                parsed = _parse_date(value)
                if not parsed:
                    return jsonify({'ok': False, 'error': 'Invalid date format. Use yyyy-mm-dd.'})
                # Validate year is reasonable
                try:
                    from datetime import datetime as dt
                    d = dt.strptime(parsed, '%Y-%m-%d')
                    if d.year < 1950 or d.year > 2099:
                        return jsonify({'ok': False, 'error': f'Year {d.year} is out of range (1950-2099).'})
                except:
                    return jsonify({'ok': False, 'error': 'Invalid date.'})
                # Check lock: don't allow moving INTO a locked period
                if lock and parsed <= lock:
                    return jsonify({'ok': False, 'error': f'Cannot set date to {parsed} — on or before lock date ({lock}). Change lock date first.'})
                # Check lock: don't allow editing a transaction that's already locked
                if lock and txn and txn['date'] <= lock:
                    return jsonify({'ok': False, 'error': f'Transaction is locked (date {txn["date"]} is before lock date {lock}).'})
                db.execute("UPDATE transactions SET date=? WHERE id=?", (parsed, txn_id))
            elif field == 'reference':
                db.execute("UPDATE transactions SET reference=? WHERE id=?", (value, txn_id))
            elif field == 'description':
                db.execute("UPDATE transactions SET description=? WHERE id=?", (value, txn_id))
                db.execute("UPDATE lines SET description=? WHERE id=?", (value, line_id))
            elif field == 'amount':
                # Check if this is a split (multi-line) transaction
                line_count = db.execute(
                    "SELECT COUNT(*) as cnt FROM lines WHERE transaction_id=?",
                    (txn_id,)).fetchone()['cnt']
                if line_count > 2:
                    return jsonify({'ok': False, 'error': 'Split transaction — edit in transaction detail view'})
                new_amt = models.parse_amount(value) if value.strip() else 0
                # The user sees amounts in normal-balance terms, so flip for credit-normal accounts
                acct_row = db.execute(
                    "SELECT a.normal_balance FROM lines l JOIN accounts a ON l.account_id=a.id WHERE l.id=?",
                    (line_id,)).fetchone()
                if acct_row and acct_row['normal_balance'] == 'C':
                    new_amt = -new_amt
                old_line = db.execute("SELECT amount, account_id FROM lines WHERE id=?", (line_id,)).fetchone()
                old_amt = old_line['amount']
                diff = new_amt - old_amt
                # Update this line
                db.execute("UPDATE lines SET amount=? WHERE id=?", (new_amt, line_id))
                # Find the cross-account line and adjust it by the opposite
                cross = db.execute(
                    "SELECT id, amount FROM lines WHERE transaction_id=? AND id!=? LIMIT 1",
                    (txn_id, line_id)).fetchone()
                if cross:
                    db.execute("UPDATE lines SET amount=? WHERE id=?", (cross['amount'] - diff, cross['id']))
            elif field == 'account':
                # Change the cross-account on a 2-line transaction
                line_count = db.execute(
                    "SELECT COUNT(*) as cnt FROM lines WHERE transaction_id=?",
                    (txn_id,)).fetchone()['cnt']
                if line_count > 2:
                    return jsonify({'ok': False, 'error': 'Split transaction — edit in transaction detail view'})
                acct_name = value.strip().upper()
                if not acct_name:
                    return jsonify({'ok': False, 'error': 'Account name required'})
                new_acct = models.get_account_by_name(acct_name)
                if not new_acct:
                    return jsonify({'ok': False, 'error': f'Account "{acct_name}" not found'})
                # Find the cross-account line (the OTHER line, not the one we're viewing)
                cross = db.execute(
                    "SELECT id, account_id FROM lines WHERE transaction_id=? AND id!=? LIMIT 1",
                    (txn_id, line_id)).fetchone()
                if cross:
                    db.execute("UPDATE lines SET account_id=? WHERE id=?", (new_acct['id'], cross['id']))
                else:
                    return jsonify({'ok': False, 'error': 'No cross-account line found'})
            else:
                return jsonify({'ok': False, 'error': f'Unknown field: {field}'})
        
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/report/<int:report_id>/sort-accounts', methods=['POST'])
def api_sort_accounts(report_id):
    """Sort account rows alphabetically within sections (between labels/totals)."""
    try:
        items = models.get_report_items(report_id)
        # Group items into sections separated by non-account rows
        sections = []
        current_section = []
        non_account_buffer = []
        
        for item in items:
            itype = item['item_type']
            if itype == 'account':
                current_section.append(item)
            else:
                if current_section:
                    sections.append(('accounts', current_section))
                    current_section = []
                sections.append(('other', [item]))
        if current_section:
            sections.append(('accounts', current_section))
        
        # Sort account sections alphabetically by name, reassign positions
        pos = 10
        with models.get_db() as db:
            for stype, sitems in sections:
                if stype == 'accounts':
                    sitems.sort(key=lambda x: (x['acct_desc'] or x['acct_name'] or '').upper())
                for item in sitems:
                    db.execute("UPDATE report_items SET position=? WHERE id=?", (pos, item['id']))
                    pos += 10
        
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/description-suggest')
def api_description_suggest():
    """Suggest descriptions based on previous entries."""
    q = request.args.get('q', '')
    if len(q) < 2:
        return jsonify([])
    with models.get_db() as db:
        rows = db.execute(
            "SELECT DISTINCT description FROM transactions WHERE description LIKE ? ORDER BY description LIMIT 15",
            (f'%{q}%',)).fetchall()
    return jsonify([r['description'] for r in rows if r['description']])

@app.route('/api/block-move', methods=['POST'])
def api_block_move():
    """Move a block of transaction lines from one account to another.
    Expects: {line_ids: [...], from_account_id: int, to_account_name: str}"""
    try:
        data = request.get_json()
        line_ids = data.get('line_ids', [])
        to_name = data.get('to_account_name', '').strip().upper()
        
        if not line_ids:
            return jsonify({'ok': False, 'error': 'No lines selected'})
        if not to_name:
            return jsonify({'ok': False, 'error': 'No target account specified'})
        
        to_acct = models.get_account_by_name(to_name)
        if not to_acct:
            return jsonify({'ok': False, 'error': f'Account "{to_name}" not found'})
        
        lock = models.get_meta('lock_date', '')
        
        with models.get_db() as db:
            moved = 0
            for lid in line_ids:
                if lock:
                    row = db.execute(
                        "SELECT t.date FROM lines l JOIN transactions t ON l.transaction_id=t.id WHERE l.id=?",
                        (lid,)).fetchone()
                    if row and row['date'] <= lock:
                        continue  # skip locked
                db.execute("UPDATE lines SET account_id=? WHERE id=?", (to_acct['id'], lid))
                moved += 1
        
        return jsonify({'ok': True, 'moved': moved})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/account-balance/<int:account_id>')
def api_account_balance(account_id):
    bal = models.get_account_balance(account_id)
    return jsonify({'balance': bal, 'formatted': models.fmt_amount(bal)})

# ─── Stripe Deposit Procedure ──────────────────────────────────────

@app.route('/api/stripe-config')
def api_stripe_config():
    """Return Stripe procedure configuration and AR customer list."""
    fee_acct = models.get_meta('stripe_fee_acct', 'EX.CC')
    ar_report_name = models.get_meta('stripe_ar_report', '')
    
    # Find AR customer accounts
    customers = []
    ar_report = None
    if ar_report_name:
        ar_report = models.find_report_by_name(ar_report_name)
    if not ar_report:
        # Try common names
        for name in ['AR', 'Accounts Receivable', 'Receivable']:
            ar_report = models.find_report_by_name(name)
            if ar_report:
                break
    if ar_report:
        accts = models.get_report_accounts(ar_report['id'])
        customers = [{'id': a['id'], 'name': a['name'], 'description': a['description']} for a in accts]
    
    # Verify fee account exists
    fee_acct_obj = models.get_account_by_name(fee_acct)
    
    return jsonify({
        'fee_account': fee_acct,
        'fee_account_ok': fee_acct_obj is not None,
        'ar_report': ar_report['name'] if ar_report else '',
        'ar_report_id': ar_report['id'] if ar_report else None,
        'customers': customers
    })

@app.route('/api/stripe-post', methods=['POST'])
def api_stripe_post():
    """Post ALL Stripe deposit charges as ONE compound distribution entry.
    
    Creates 1 transaction structured as:
      Line 0:  DR clearing_account  (total net — this is what shows in the clearing ledger)
      Then for each charge:
        CR customer_account  (gross amount)
        DR fee_account       (fee amount)  — omitted if fee is 0
    
    The clearing ledger shows ONE row for the total net deposit.
    F3/double-click opens the distribution showing all the customer + fee detail.
    
    REF format: S.MMMdd (e.g. S.Jun01)
    Description: Stripe deposits, net
    """
    try:
        from datetime import datetime as dt
        data = request.get_json()
        deposit_date = data.get('date', '')
        items = data.get('items', [])
        clearing_account_id = data.get('clearing_account_id')
        fee_acct_name = data.get('fee_account', 'EX.CC')
        
        if not deposit_date:
            return jsonify({'ok': False, 'error': 'Deposit date is required'})
        parsed_date = _parse_date(deposit_date)
        if not parsed_date:
            return jsonify({'ok': False, 'error': 'Invalid date format'})
        
        if not items:
            return jsonify({'ok': False, 'error': 'No line items to post'})
        
        # Build REF as S.MMMdd
        try:
            d = dt.strptime(parsed_date, '%Y-%m-%d')
            ref = 'S.' + d.strftime('%b%d')  # e.g. S.Jun01
        except:
            ref = 'S.Stripe'
        
        # Resolve fee account
        fee_acct = models.get_account_by_name(fee_acct_name)
        if not fee_acct:
            return jsonify({'ok': False, 'error': f'Fee account "{fee_acct_name}" not found'})
        
        # Resolve clearing account
        if not clearing_account_id:
            return jsonify({'ok': False, 'error': 'Clearing account not specified'})
        clearing_acct = models.get_account(clearing_account_id)
        if not clearing_acct:
            return jsonify({'ok': False, 'error': 'Clearing account not found'})
        
        # Validate all lines first, accumulate totals and build detail lines
        detail_lines = []  # customer CRs and fee DRs
        total_net = 0
        total_fees = 0
        total_gross = 0
        errors = []
        
        for i, item in enumerate(items):
            customer_name = item.get('customer', '').strip().upper()
            gross_str = item.get('gross', '0')
            fee_str = item.get('fee', '0')
            description = item.get('description', '').strip()
            
            if not customer_name:
                errors.append(f'Line {i+1}: Client AR account required')
                continue
            
            gross_cents = models.parse_amount(gross_str)
            fee_cents = models.parse_amount(fee_str)
            
            if gross_cents <= 0:
                errors.append(f'Line {i+1}: Gross must be positive')
                continue
            if fee_cents < 0:
                errors.append(f'Line {i+1}: Fee cannot be negative')
                continue
            
            net_cents = gross_cents - fee_cents
            
            cust_acct = models.get_account_by_name(customer_name)
            if not cust_acct:
                errors.append(f'Line {i+1}: Account "{customer_name}" not found')
                continue
            
            line_desc = description or 'Stripe pmt'
            
            # CR customer AR (gross) — goes into distribution detail
            detail_lines.append((cust_acct['id'], -gross_cents, line_desc))
            # DR fee account — goes into distribution detail
            if fee_cents > 0:
                detail_lines.append((fee_acct['id'], fee_cents, line_desc))
            
            total_net += net_cents
            total_fees += fee_cents
            total_gross += gross_cents
        
        if errors:
            return jsonify({'ok': False, 'error': 'Validation errors', 'errors': errors})
        
        if not detail_lines:
            return jsonify({'ok': False, 'error': 'No valid lines to post'})
        
        # Build the full transaction:
        # Line 0: ONE DR to clearing for the total net (this is the single ledger row)
        # Lines 1+: all the customer CRs and fee DRs (distribution detail via F3)
        txn_desc = 'Stripe deposits, net'
        all_lines = [(clearing_acct['id'], total_net, txn_desc)] + detail_lines
        
        models.add_transaction(parsed_date, ref, txn_desc, all_lines)
        
        return jsonify({
            'ok': True,
            'posted': len(items),
            'total_items': len(items),
            'total_net': models.fmt_amount(total_net),
            'total_fees': models.fmt_amount(total_fees),
            'total_gross': models.fmt_amount(total_gross),
            'errors': []
        })
        
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ─── CSV Export ─────────────────────────────────────────────────────

@app.route('/export/ledger/<int:account_id>')
def export_ledger(account_id):
    import csv, io
    account = models.get_account(account_id)
    entries = models.get_ledger(account_id)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date', 'Reference', 'Description', 'Cross Account', 
                    'Debit', 'Credit', 'Balance'])
    for e in entries:
        dr = models.fmt_amount_plain(e['amount']) if e['amount'] > 0 else ''
        cr = models.fmt_amount_plain(abs(e['amount'])) if e['amount'] < 0 else ''
        writer.writerow([e['date'], e['reference'], e['description'],
                        e['cross_accounts'], dr, cr,
                        models.fmt_amount_plain(e['running_balance'])])
    output.seek(0)
    return output.getvalue(), 200, {
        'Content-Type': 'text/csv',
        'Content-Disposition': f'attachment; filename={account["name"]}_ledger.csv'
    }

@app.route('/export/trial-balance')
def export_trial_balance():
    import csv, io
    accounts, total_dr, total_cr = models.get_trial_balance()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Account', 'Number', 'Description', 'Debit', 'Credit'])
    for a in accounts:
        writer.writerow([a['name'], a['account_number'], a['description'],
                        models.fmt_amount_plain(a['debit']) if a['debit'] else '',
                        models.fmt_amount_plain(a['credit']) if a['credit'] else ''])
    writer.writerow(['', '', 'TOTALS', models.fmt_amount_plain(total_dr),
                    models.fmt_amount_plain(total_cr)])
    output.seek(0)
    return output.getvalue(), 200, {
        'Content-Type': 'text/csv',
        'Content-Disposition': 'attachment; filename=trial_balance.csv'
    }

# ─── Books Export / Import (Full Backup & Restore) ───────────────

@app.route('/export/structure')
def export_structure():
    """Export the full account framework: meta, accounts, reports+items, tax codes, import rules."""
    if not models.get_db_path():
        return redirect(url_for('library'))

    company = models.get_meta('company_name', 'Books')

    with models.get_db() as db:
        # Meta (all settings)
        meta = {}
        for row in db.execute("SELECT key, value FROM meta").fetchall():
            meta[row['key']] = row['value']

        # Accounts — keyed by name for portability
        accounts = []
        for a in db.execute("SELECT * FROM accounts ORDER BY id").fetchall():
            accounts.append({
                'name': a['name'],
                'description': a['description'],
                'normal_balance': a['normal_balance'],
                'account_type': a['account_type'],
                'account_number': a['account_number'] or '',
                'notes': a['notes'] or '',
            })

        # Reports with their items
        reports = []
        for r in db.execute("SELECT * FROM reports ORDER BY sort_order, id").fetchall():
            items = []
            for it in db.execute(
                "SELECT ri.*, a.name as acct_name FROM report_items ri "
                "LEFT JOIN accounts a ON ri.account_id = a.id "
                "WHERE ri.report_id=? ORDER BY ri.position", (r['id'],)).fetchall():
                items.append({
                    'position': it['position'],
                    'item_type': it['item_type'],
                    'description': it['description'] or '',
                    'account_name': it['acct_name'] or '',
                    'indent': it['indent'],
                    'total_to_1': it['total_to_1'] or '',
                    'total_to_2': it['total_to_2'] or '',
                    'total_to_3': it['total_to_3'] or '',
                    'total_to_4': it['total_to_4'] or '',
                    'total_to_5': it['total_to_5'] or '',
                    'total_to_6': it['total_to_6'] or '',
                    'sep_style': it['sep_style'] or '',
                })
            reports.append({
                'name': r['name'],
                'description': r['description'] or '',
                'sort_order': r['sort_order'],
                'period_begin': r['period_begin'] or '',
                'period_end': r['period_end'] or '',
                'items': items,
            })

        # Tax codes
        tax_codes = []
        for tc in db.execute("SELECT * FROM tax_codes ORDER BY id").fetchall():
            tax_codes.append({
                'id': tc['id'],
                'description': tc['description'] or '',
                'rate_percent': tc['rate_percent'],
                'collected_account': tc['collected_account'] or '',
                'paid_account': tc['paid_account'] or '',
            })

        # Import rules
        import_rules = []
        for ir in db.execute("SELECT * FROM import_rules ORDER BY priority DESC, keyword").fetchall():
            import_rules.append({
                'keyword': ir['keyword'],
                'account_name': ir['account_name'],
                'tax_code': ir['tax_code'] or '',
                'priority': ir['priority'],
                'notes': ir['notes'] or '',
            })

    payload = {
        '_grid_export': 'structure',
        '_version': 1,
        '_exported': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        '_company': company,
        'meta': meta,
        'accounts': accounts,
        'reports': reports,
        'tax_codes': tax_codes,
        'import_rules': import_rules,
    }

    data = json.dumps(payload, indent=2)
    safe_name = company.replace(' ', '_').replace('/', '-')[:40]
    return data, 200, {
        'Content-Type': 'application/json',
        'Content-Disposition': f'attachment; filename={safe_name}_structure.json'
    }


@app.route('/export/data')
def export_data():
    """Export all transactions and lines. Accounts referenced by name."""
    if not models.get_db_path():
        return redirect(url_for('library'))

    company = models.get_meta('company_name', 'Books')

    with models.get_db() as db:
        # Build account ID → name map
        acct_map = {}
        for a in db.execute("SELECT id, name FROM accounts").fetchall():
            acct_map[a['id']] = a['name']

        transactions = []
        for t in db.execute("SELECT * FROM transactions ORDER BY date, id").fetchall():
            lines = []
            for ln in db.execute(
                "SELECT * FROM lines WHERE transaction_id=? ORDER BY sort_order", (t['id'],)).fetchall():
                lines.append({
                    'account_name': acct_map.get(ln['account_id'], f'UNKNOWN_{ln["account_id"]}'),
                    'amount': ln['amount'],
                    'description': ln['description'] or '',
                    'reconciled': ln['reconciled'],
                    'sort_order': ln['sort_order'],
                })
            transactions.append({
                'date': t['date'],
                'reference': t['reference'] or '',
                'description': t['description'] or '',
                'created_at': t['created_at'] or '',
                'lines': lines,
            })

    payload = {
        '_grid_export': 'data',
        '_version': 1,
        '_exported': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        '_company': company,
        '_transaction_count': len(transactions),
        'transactions': transactions,
    }

    data = json.dumps(payload, indent=2)
    safe_name = company.replace(' ', '_').replace('/', '-')[:40]
    return data, 200, {
        'Content-Type': 'application/json',
        'Content-Disposition': f'attachment; filename={safe_name}_data.json'
    }


@app.route('/import/structure', methods=['POST'])
def import_structure():
    """Import structure JSON into the CURRENT open books file.
    WARNING: This replaces all accounts, reports, tax codes, and rules."""
    if not models.get_db_path():
        flash('No books file open', 'error')
        return redirect(url_for('library'))

    f = request.files.get('file')
    if not f:
        flash('No file selected', 'error')
        return redirect(url_for('settings'))

    try:
        raw = f.read().decode('utf-8')
        payload = json.loads(raw)
    except Exception as e:
        flash(f'Invalid JSON file: {e}', 'error')
        return redirect(url_for('settings'))

    if payload.get('_grid_export') != 'structure':
        flash('This is not a Grid structure export file', 'error')
        return redirect(url_for('settings'))

    # Check if there are existing transactions — warn user
    with models.get_db() as db:
        txn_count = db.execute("SELECT COUNT(*) as cnt FROM transactions").fetchone()['cnt']
        if txn_count > 0 and not request.form.get('confirm_overwrite'):
            flash(f'This file has {txn_count} transactions. Import structure into a FRESH file '
                  '(create one first via Library → New). Importing structure wipes accounts/reports.', 'error')
            return redirect(url_for('settings'))

    try:
        with models.get_db() as db:
            # Clear existing structure (order matters for FK constraints)
            db.execute("DELETE FROM report_items")
            db.execute("DELETE FROM reports")
            # Only delete accounts that have no transaction lines
            # For a fresh file this deletes all; for a file with data, preserve referenced accounts
            if txn_count == 0:
                db.execute("DELETE FROM accounts")
            db.execute("DELETE FROM tax_codes")
            db.execute("DELETE FROM import_rules")

            # Meta
            for key, value in payload.get('meta', {}).items():
                db.execute("INSERT OR REPLACE INTO meta(key, value) VALUES(?,?)", (key, value))

            # Accounts
            acct_name_to_id = {}
            for a in payload.get('accounts', []):
                # Check if account already exists (e.g. in a file with data)
                existing = db.execute("SELECT id FROM accounts WHERE name=? COLLATE NOCASE",
                                     (a['name'],)).fetchone()
                if existing:
                    acct_name_to_id[a['name']] = existing['id']
                    # Update metadata
                    db.execute("UPDATE accounts SET description=?, normal_balance=?, account_type=?, "
                              "account_number=?, notes=? WHERE id=?",
                              (a.get('description', ''), a['normal_balance'], a['account_type'],
                               a.get('account_number', ''), a.get('notes', ''), existing['id']))
                else:
                    cur = db.execute(
                        "INSERT INTO accounts(name, description, normal_balance, account_type, account_number, notes) "
                        "VALUES(?,?,?,?,?,?)",
                        (a['name'], a.get('description', ''), a['normal_balance'],
                         a['account_type'], a.get('account_number', ''), a.get('notes', '')))
                    acct_name_to_id[a['name']] = cur.lastrowid

            # Reports + items
            for r in payload.get('reports', []):
                cur = db.execute(
                    "INSERT INTO reports(name, description, sort_order, period_begin, period_end) VALUES(?,?,?,?,?)",
                    (r['name'], r.get('description', ''), r.get('sort_order', 0),
                     r.get('period_begin', ''), r.get('period_end', '')))
                report_id = cur.lastrowid
                for it in r.get('items', []):
                    acct_id = acct_name_to_id.get(it.get('account_name')) if it.get('account_name') else None
                    db.execute(
                        "INSERT INTO report_items(report_id, position, item_type, description, account_id, "
                        "indent, total_to_1, total_to_2, total_to_3, total_to_4, total_to_5, total_to_6, sep_style) "
                        "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (report_id, it.get('position', 0), it['item_type'], it.get('description', ''),
                         acct_id, it.get('indent', 0),
                         it.get('total_to_1', ''), it.get('total_to_2', ''), it.get('total_to_3', ''),
                         it.get('total_to_4', ''), it.get('total_to_5', ''), it.get('total_to_6', ''),
                         it.get('sep_style', '')))

            # Tax codes
            for tc in payload.get('tax_codes', []):
                db.execute(
                    "INSERT OR REPLACE INTO tax_codes(id, description, rate_percent, collected_account, paid_account) "
                    "VALUES(?,?,?,?,?)",
                    (tc['id'], tc.get('description', ''), tc.get('rate_percent', 0),
                     tc.get('collected_account', ''), tc.get('paid_account', '')))

            # Import rules
            for ir in payload.get('import_rules', []):
                db.execute(
                    "INSERT INTO import_rules(keyword, account_name, tax_code, priority, notes) VALUES(?,?,?,?,?)",
                    (ir['keyword'], ir['account_name'], ir.get('tax_code', ''),
                     ir.get('priority', 0), ir.get('notes', '')))

        acct_count = len(payload.get('accounts', []))
        rpt_count = len(payload.get('reports', []))
        rule_count = len(payload.get('import_rules', []))
        flash(f'Structure imported: {acct_count} accounts, {rpt_count} reports, {rule_count} rules', 'success')

    except Exception as e:
        flash(f'Import failed: {e}', 'error')

    return redirect(url_for('settings'))


@app.route('/import/data', methods=['POST'])
def import_data():
    """Import transactions from a data JSON file into the current books."""
    if not models.get_db_path():
        flash('No books file open', 'error')
        return redirect(url_for('library'))

    f = request.files.get('file')
    if not f:
        flash('No file selected', 'error')
        return redirect(url_for('settings'))

    try:
        raw = f.read().decode('utf-8')
        payload = json.loads(raw)
    except Exception as e:
        flash(f'Invalid JSON file: {e}', 'error')
        return redirect(url_for('settings'))

    if payload.get('_grid_export') != 'data':
        flash('This is not a Grid data export file', 'error')
        return redirect(url_for('settings'))

    # Build account name → id map from current database
    with models.get_db() as db:
        acct_map = {}
        for a in db.execute("SELECT id, name FROM accounts").fetchall():
            acct_map[a['name'].upper()] = a['id']

    # Temporarily disable lock date for import
    saved_lock = models.get_meta('lock_date', '')
    if saved_lock:
        models.set_meta('lock_date', '')

    imported = 0
    skipped = 0
    errors = []

    try:
        with models.get_db() as db:
            for txn in payload.get('transactions', []):
                # Resolve account names to IDs
                lines_resolved = []
                missing = []
                for ln in txn.get('lines', []):
                    acct_name = ln['account_name']
                    acct_id = acct_map.get(acct_name.upper())
                    if not acct_id:
                        missing.append(acct_name)
                    else:
                        lines_resolved.append((acct_id, ln['amount'], ln.get('description', ''),
                                              ln.get('reconciled', 0), ln.get('sort_order', 0)))

                if missing:
                    errors.append(f"Txn {txn['date']} '{txn['description']}': missing accounts: {', '.join(missing)}")
                    skipped += 1
                    continue

                # Check balance
                total = sum(lr[1] for lr in lines_resolved)
                if total != 0:
                    errors.append(f"Txn {txn['date']} '{txn['description']}': does not balance (off by {total})")
                    skipped += 1
                    continue

                # Insert transaction
                cur = db.execute(
                    "INSERT INTO transactions(date, reference, description, created_at) VALUES(?,?,?,?)",
                    (txn['date'], txn.get('reference', ''), txn.get('description', ''),
                     txn.get('created_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))))
                txn_id = cur.lastrowid

                # Insert lines with reconciled flag
                for acct_id, amount, desc, reconciled, sort_order in lines_resolved:
                    db.execute(
                        "INSERT INTO lines(transaction_id, account_id, amount, description, reconciled, sort_order) "
                        "VALUES(?,?,?,?,?,?)",
                        (txn_id, acct_id, amount, desc, reconciled, sort_order))

                imported += 1

    except Exception as e:
        flash(f'Import failed at transaction {imported + 1}: {e}', 'error')
        # Restore lock date
        if saved_lock:
            models.set_meta('lock_date', saved_lock)
        return redirect(url_for('settings'))

    # Restore lock date
    if saved_lock:
        models.set_meta('lock_date', saved_lock)

    if errors:
        for err in errors[:10]:  # Show first 10 errors
            flash(err, 'error')
        if len(errors) > 10:
            flash(f'... and {len(errors) - 10} more errors', 'error')

    flash(f'Data imported: {imported} transactions ({skipped} skipped)', 'success')
    return redirect(url_for('settings'))


# ─── Import Rules Management ──────────────────────────────────────

@app.route('/rules', methods=['GET'])
def import_rules_page():
    if not models.get_db_path():
        return redirect(url_for('library'))
    company = models.get_meta('company_name', 'My Books')
    rules = models.get_import_rules()
    tax_codes = models.get_tax_codes()
    accounts = models.get_accounts()
    return render_template('rules.html', company=company, rules=rules,
        tax_codes=tax_codes, accounts=accounts)

@app.route('/api/rule-add', methods=['POST'])
def api_rule_add():
    """AJAX endpoint: add an import rule from the import preview screen."""
    try:
        data = request.get_json()
        keyword = (data.get('keyword', '') or '').strip()
        account_name = (data.get('account_name', '') or '').strip().upper()
        tax_code = (data.get('tax_code', '') or '').strip()
        priority = int(data.get('priority', 5))
        if not keyword or not account_name:
            return jsonify({'ok': False, 'error': 'Keyword and account are required'})
        # Verify account exists
        acct = models.get_account_by_name(account_name)
        if not acct:
            return jsonify({'ok': False, 'error': f'Account "{account_name}" not found'})
        models.save_import_rule(None, keyword, account_name, tax_code, priority, '')
        return jsonify({'ok': True, 'keyword': keyword, 'account': account_name, 'tax': tax_code})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/rules/save', methods=['POST'])
def save_rule():
    rid = request.form.get('rule_id', '')
    keyword = request.form.get('keyword', '').strip()
    account_name = request.form.get('account_name', '').strip().upper()
    tax_code = request.form.get('tax_code', '').strip()
    priority = int(request.form.get('priority', '0'))
    notes = request.form.get('notes', '').strip()
    if keyword and account_name:
        models.save_import_rule(int(rid) if rid else None, keyword, account_name, tax_code, priority, notes)
        flash(f'Rule saved: "{keyword}" → {account_name}', 'success')
    else:
        flash('Keyword and account are required', 'error')
    return redirect(url_for('import_rules_page'))

@app.route('/rules/delete/<int:rule_id>', methods=['POST'])
def delete_rule(rule_id):
    models.delete_import_rule(rule_id)
    flash('Rule deleted', 'success')
    return redirect(url_for('import_rules_page'))

@app.route('/tax/save', methods=['POST'])
def save_tax():
    code_id = request.form.get('code_id', '').strip().upper()
    description = request.form.get('description', '').strip()
    rate = float(request.form.get('rate_percent', '0'))
    collected = request.form.get('collected_account', '').strip()
    paid = request.form.get('paid_account', '').strip()
    if code_id:
        models.save_tax_code(code_id, description, rate, collected, paid)
        flash(f'Tax code {code_id} saved', 'success')
    return redirect(url_for('import_rules_page'))

@app.route('/tax/delete/<code_id>', methods=['POST'])
def delete_tax(code_id):
    models.delete_tax_code(code_id)
    flash(f'Tax code {code_id} deleted', 'success')
    return redirect(url_for('import_rules_page'))

# ─── Rules Export/Import ─────────────────────────────────────────

@app.route('/rules/export')
def export_rules():
    """Export import rules as CSV."""
    import csv, io
    rules = models.get_import_rules() if hasattr(models, 'get_import_rules') else []
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['keyword', 'account', 'description', 'tax_code', 'priority'])
    for r in rules:
        w.writerow([r['keyword'], r['account_name'], r['notes'] or '',
                    r['tax_code'] or '', r['priority'] or 0])
    from flask import Response
    return Response(buf.getvalue(), mimetype='text/csv',
                   headers={'Content-Disposition': 'attachment; filename=rules.csv'})

@app.route('/rules/import', methods=['POST'])
def import_rules_csv():
    """Import rules from CSV file."""
    import csv, io
    f = request.files.get('file')
    if not f:
        flash('No file selected', 'error')
        return redirect(url_for('import_rules_page'))
    text = f.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    count = 0
    for row in reader:
        kw = row.get('keyword', '').strip()
        acct = row.get('account', '').strip()
        desc = row.get('description', '').strip()
        tax = row.get('tax_code', '').strip()
        pri = int(row.get('priority', '0') or '0')
        if kw and acct:
            models.save_import_rule(None, kw, acct, tax, pri, desc)
            count += 1
    flash(f'Imported {count} rules', 'success')
    return redirect(url_for('import_rules_page'))

# ─── Template Clone (New File from Existing) ─────────────────────

@app.route('/clone', methods=['GET', 'POST'])
def clone_file():
    """Create a new client file from an existing one — keeps COA, reports, rules; purges transactions."""
    if request.method == 'POST':
        source = request.form.get('source', '')
        new_company = request.form.get('company', '').strip()
        new_folder = request.form.get('folder', '').strip()
        if not source or not new_company or not new_folder:
            flash('All fields required', 'error')
            return redirect(url_for('clone_file'))
        
        import shutil
        base = os.path.dirname(os.path.dirname(source))
        new_dir = os.path.join(base, new_folder)
        os.makedirs(new_dir, exist_ok=True)
        new_path = os.path.join(new_dir, 'books.db')
        if os.path.exists(new_path):
            flash('File already exists in that folder', 'error')
            return redirect(url_for('clone_file'))
        
        shutil.copy2(source, new_path)
        
        # Purge transactions from the new file
        import sqlite3
        conn = sqlite3.connect(new_path)
        conn.execute("DELETE FROM lines")
        conn.execute("DELETE FROM transactions")
        conn.execute("UPDATE meta SET value=? WHERE key='company_name'", (new_company,))
        # Reset any column configs
        conn.execute("DELETE FROM meta WHERE key LIKE 'columns_%'")
        conn.commit()
        conn.close()
        
        flash(f'Created new file for "{new_company}" in {new_folder}/', 'success')
        return redirect(url_for('library'))
    
    # GET: show form
    clients = list_client_books()
    company = models.get_meta('company_name', '') if models.get_db_path() else ''
    return render_template('clone.html', clients=clients, company=company)

# ─── Reports Section ─────────────────────────────────────────────

@app.route('/reports')
def reports_page():
    """Reports landing page."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    company = models.get_meta('company_name', 'My Books')
    fiscal_ye = models.get_meta('fiscal_year_end', '12-31')
    # Compute default fiscal year dates
    from datetime import date, timedelta
    today = date.today()
    fy_month, fy_day = [int(x) for x in fiscal_ye.split('-')]
    # Current fiscal year end
    fy_end_date = date(today.year, fy_month, fy_day)
    if fy_end_date < today:
        # FYE already passed this year, use next year
        fy_end_date = date(today.year + 1, fy_month, fy_day)
    # Fiscal year begin is day after previous FYE
    prev_fye = date(fy_end_date.year - 1, fy_month, fy_day)
    fy_begin_date = prev_fye + timedelta(days=1)
    
    return render_template('reports.html', company=company,
        fiscal_display=fy_end_date.strftime('%d %b %Y'),
        fy_begin=fy_begin_date.strftime('%Y-%m-%d'),
        fy_end=fy_end_date.strftime('%Y-%m-%d'))


def _get_bs_account_ids():
    """Return set of account_ids that appear in BS report."""
    reports = models.get_reports()
    bs = next((r for r in reports if r['name'] == 'BS'), None)
    if not bs:
        return set()
    items = models.get_report_items(bs['id'])
    return {i['account_id'] for i in items if i['account_id']}


def _get_report_account_order(report_name):
    """Return ordered list of (account_id, acct_name, acct_desc) from a report's items."""
    reports = models.get_reports()
    rpt = next((r for r in reports if r['name'] == report_name), None)
    if not rpt:
        return []
    items = models.get_report_items(rpt['id'])
    seen = set()
    result = []
    for item in items:
        aid = item['account_id']
        atype = item['account_type'] or ''
        if aid and aid not in seen and atype == 'posting':
            seen.add(aid)
            result.append((aid, item['acct_name'] or '', item['acct_desc'] or ''))
    return result


def _build_account_detail(account_id, acct_name, acct_desc, begin, end, is_bs, dr_cr_filter='all'):
    """Build GL detail rows for one account.
    Returns list of dicts: {type, date, ref, desc, debit, credit, balance}
    Debits positive, credits negative. No normal-balance flipping.
    """
    rows = []
    
    # Opening balance
    if is_bs:
        # Sum all transactions up to day before begin
        from datetime import datetime, timedelta
        if begin:
            d = datetime.strptime(begin, '%Y-%m-%d') - timedelta(days=1)
            opening = models.get_account_balance(account_id, date_to=d.strftime('%Y-%m-%d'))
        else:
            opening = 0
    else:
        opening = 0
    
    # Get transactions in period
    with models.get_db() as db:
        sql = """
            SELECT t.id as txn_id, t.date, t.reference, t.description as txn_desc,
                   l.amount, l.description as line_desc, l.id as line_id,
                   GROUP_CONCAT(DISTINCT a2.name) as cross_accounts
            FROM lines l
            JOIN transactions t ON l.transaction_id = t.id
            LEFT JOIN lines l2 ON l2.transaction_id = t.id AND l2.account_id != ?
            LEFT JOIN accounts a2 ON l2.account_id = a2.id
            WHERE l.account_id = ?"""
        params = [account_id, account_id]
        if begin: sql += " AND t.date >= ?"; params.append(begin)
        if end: sql += " AND t.date <= ?"; params.append(end)
        sql += " GROUP BY l.id ORDER BY t.date, t.id, l.sort_order"
        txns = db.execute(sql, params).fetchall()
    
    balance = opening
    
    for txn in txns:
        amt = txn['amount']  # positive = debit, negative = credit
        debit = amt if amt > 0 else 0
        credit = -amt if amt < 0 else 0
        
        if dr_cr_filter == 'debit' and amt <= 0:
            continue
        if dr_cr_filter == 'credit' and amt >= 0:
            continue
        
        balance += amt
        cross_raw = txn['cross_accounts'] or ''
        cross = '-split-' if ',' in cross_raw else cross_raw
        
        rows.append({
            'type': 'txn',
            'date': txn['date'],
            'ref': txn['reference'] or '',
            'desc': txn['line_desc'] or txn['txn_desc'] or '',
            'debit': debit,
            'credit': credit,
            'balance': balance,
            'cross': cross,
        })
    
    closing = balance
    return opening, rows, closing


def _fmt_money(cents):
    """Format cents as dollar string. No normal-balance. Debits positive."""
    if cents == 0:
        return '—'
    neg = cents < 0
    val = abs(cents) / 100.0
    s = f'{val:,.2f}'
    return f'({s})' if neg else s


@app.route('/reports/gl')
def report_gl():
    """Generate General Ledger report."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    
    begin = request.args.get('begin', '')
    end = request.args.get('end', '')
    fmt = request.args.get('fmt', 'pdf')
    dr_cr_filter = request.args.get('filter', 'all')
    company = models.get_meta('company_name', 'My Books')
    
    bs_ids = _get_bs_account_ids()
    
    # Get accounts in BS order then IS order
    accounts = _get_report_account_order('BS') + _get_report_account_order('IS')
    
    if fmt == 'csv':
        return _gl_csv(accounts, bs_ids, begin, end, dr_cr_filter, company)
    else:
        try:
            return _gl_pdf(accounts, bs_ids, begin, end, dr_cr_filter, company)
        except Exception as e:
            flash(f'PDF error: {e}. Install reportlab: pip install reportlab', 'error')
            return redirect(url_for('reports_page'))


def _gl_csv(accounts, bs_ids, begin, end, dr_cr_filter, company):
    """Generate GL as CSV download."""
    import csv, io
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(['Account', 'Description', 'Date', 'Reference', 'Detail', 'Debit', 'Credit', 'Balance', 'Cross Account'])
    
    for aid, aname, adesc in accounts:
        is_bs = aid in bs_ids
        opening, rows, closing = _build_account_detail(aid, aname, adesc, begin, end, is_bs, dr_cr_filter)
        
        if not rows and opening == 0:
            continue  # skip empty accounts
        
        # Opening balance row
        w.writerow([aname, adesc, begin or '', '', 'Opening Balance',
                    _fmt_money(opening) if opening > 0 else '',
                    _fmt_money(-opening) if opening < 0 else '',
                    _fmt_money(opening), ''])
        
        for r in rows:
            w.writerow([aname, adesc, r['date'], r['ref'], r['desc'],
                        _fmt_money(r['debit']) if r['debit'] else '',
                        _fmt_money(r['credit']) if r['credit'] else '',
                        _fmt_money(r['balance']), r['cross']])
        
        # Closing balance row
        w.writerow([aname, adesc, end or '', '', 'Closing Balance', '', '', _fmt_money(closing), ''])
        w.writerow([])  # blank row between accounts
    
    resp = app.make_response(output.getvalue())
    resp.headers['Content-Type'] = 'text/csv'
    resp.headers['Content-Disposition'] = f'attachment; filename=GL_{begin}_{end}.csv'
    return resp


def _gl_pdf(accounts, bs_ids, begin, end, dr_cr_filter, company):
    """Generate GL as monospaced PDF."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        from flask import abort
        abort(500, 'reportlab not installed. Run: pip install reportlab')
    import io, os
    
    # Try monospaced fonts in preference order, use Courier as final fallback
    font = 'Courier'
    font_b = 'Courier-Bold'
    candidates = [
        # (reg_name, regular_path, bold_path) — checked in order
        # Linux
        ('LiberationMono', '/usr/share/fonts/truetype/liberation/LiberationMono-Regular.ttf',
                           '/usr/share/fonts/truetype/liberation/LiberationMono-Bold.ttf'),
        ('DejaVuMono',     '/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf',
                           '/usr/share/fonts/truetype/dejavu/DejaVuSansMono-Bold.ttf'),
        # Windows 10+
        ('Consolas',       'C:/Windows/Fonts/consola.ttf',
                           'C:/Windows/Fonts/consolab.ttf'),
        # Windows 7+
        ('CourierNew',     'C:/Windows/Fonts/cour.ttf',
                           'C:/Windows/Fonts/courbd.ttf'),
        # macOS
        ('Menlo',          '/System/Library/Fonts/Menlo.ttc',
                           '/System/Library/Fonts/Menlo.ttc'),
    ]
    for name, regular, bold in candidates:
        if os.path.exists(regular) and os.path.exists(bold):
            try:
                # Only register if not already registered
                try: pdfmetrics.getFont(name)
                except KeyError:
                    pdfmetrics.registerFont(TTFont(name, regular))
                    pdfmetrics.registerFont(TTFont(name+'-Bold', bold))
                font = name
                font_b = name + '-Bold'
                break
            except Exception:
                continue
    
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    pw, ph = letter  # 612, 792
    margin = 36  # 0.5 inch all sides
    usable_w = pw - 2 * margin
    
    fs = 6.5  # tighter font for more description space
    line_h = 8.5
    
    # Column layout within 0.5" margins (usable = 540pt)
    # Tightened: date 40pt, ref 26pt, then desc gets the rest up to debit
    col_date = margin
    col_ref = margin + 40
    col_desc = margin + 66
    col_debit = margin + 310
    col_credit = margin + 375
    col_balance = margin + 440
    col_cross = margin + 510
    right_edge = pw - margin
    desc_max = 62  # chars that fit between col_desc and col_debit at 6.5pt
    
    y = ph - margin
    page_num = 1
    
    def short_date(d):
        """Format date as dd-Mon-yy"""
        if not d: return ''
        from datetime import datetime
        try:
            dt = datetime.strptime(d[:10], '%Y-%m-%d')
            return dt.strftime('%d-%b-%y')
        except: return d[:10]
    
    begin_s = short_date(begin) if begin else 'Start'
    end_s = short_date(end) if end else 'Current'
    
    def header():
        nonlocal y
        c.setFont(font_b, 8)
        c.drawString(margin, ph - margin + 5, f'{company} — General Ledger')
        c.setFont(font, 6)
        c.drawString(margin, ph - margin - 4, f'{begin_s} to {end_s}')
        c.drawRightString(right_edge, ph - margin + 5, f'Page {page_num}')
        y = ph - margin - 12
    
    def col_header():
        nonlocal y
        c.setFont(font_b, fs)
        c.drawString(col_date, y, 'Date')
        c.drawString(col_ref, y, 'Ref')
        c.drawString(col_desc, y, 'Description')
        c.drawRightString(col_debit + 58, y, 'Debit')
        c.drawRightString(col_credit + 58, y, 'Credit')
        c.drawRightString(col_balance + 58, y, 'Balance')
        c.drawString(col_cross, y, 'Acct')
        y -= 2
        c.setLineWidth(0.4)
        c.line(margin, y, right_edge, y)
        y -= line_h
    
    def check_page(need=2):
        nonlocal y, page_num
        if y < margin + need * line_h:
            c.showPage()
            page_num += 1
            header()
            col_header()
    
    def draw_row(date_s, ref_s, desc_s, debit, credit, balance, cross='', bold=False):
        nonlocal y
        check_page()
        fn = font_b if bold else font
        c.setFont(fn, fs)
        c.drawString(col_date, y, date_s)
        c.drawString(col_ref, y, (ref_s or '')[:6])
        c.drawString(col_desc, y, (desc_s or '')[:desc_max])
        if debit: c.drawRightString(col_debit + 58, y, _fmt_money(debit))
        # Credits always in brackets
        if credit: c.drawRightString(col_credit + 58, y, _fmt_money(-credit))
        if balance is not None: c.drawRightString(col_balance + 58, y, _fmt_money(balance))
        if cross: c.drawString(col_cross, y, cross[:12])
        y -= line_h
    
    header()
    
    for idx, (aid, aname, adesc) in enumerate(accounts):
        is_bs = aid in bs_ids
        opening, rows, closing = _build_account_detail(aid, aname, adesc, begin, end, is_bs, dr_cr_filter)
        
        if not rows and opening == 0:
            continue
        
        check_page(5)
        
        # Account header
        c.setFont(font_b, 8)
        c.drawString(margin, y, f'{aname}  {adesc}')
        y -= line_h
        col_header()
        
        # Opening balance — only in balance column
        draw_row(begin_s, '', 'Opening Balance', 0, 0, opening, bold=True)
        
        # Transaction rows — debit positive, credit in brackets, always
        total_dr, total_cr = 0, 0
        for r in rows:
            total_dr += r['debit']
            total_cr += r['credit']
            draw_row(short_date(r['date']), r['ref'], r['desc'],
                     r['debit'], r['credit'], r['balance'], r['cross'])
        
        # Single underline before closing
        check_page()
        c.setLineWidth(0.3)
        c.line(col_debit, y + line_h - 2, col_balance + 68, y + line_h - 2)
        
        # Closing line: debit/credit show TOTALS, balance shows closing balance
        draw_row(end_s, '', 'Closing Balance', total_dr, total_cr, closing, bold=True)
        
        # Double underline on balance column
        c.setLineWidth(0.4)
        c.line(col_balance, y + line_h - 2, col_balance + 68, y + line_h - 2)
        c.line(col_balance, y + line_h - 5, col_balance + 68, y + line_h - 5)
        y -= line_h * 0.5  # half-line gap between accounts
    
    c.save()
    buf.seek(0)
    
    resp = app.make_response(buf.read())
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'inline; filename=GL_{begin}_{end}.pdf'
    return resp


@app.route('/reports/account')
def report_account_detail():
    """Generate single account detail report."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    
    acct_name = request.args.get('account', '').strip().upper()
    begin = request.args.get('begin', '')
    end = request.args.get('end', '')
    fmt = request.args.get('fmt', 'pdf')
    dr_cr_filter = request.args.get('filter', 'all')
    company = models.get_meta('company_name', 'My Books')
    
    acct = models.get_account_by_name(acct_name)
    if not acct:
        flash(f'Account "{acct_name}" not found', 'error')
        return redirect(url_for('reports_page'))
    
    bs_ids = _get_bs_account_ids()
    accounts = [(acct['id'], acct['name'], acct['description'])]
    
    if fmt == 'csv':
        return _gl_csv(accounts, bs_ids, begin, end, dr_cr_filter, company)
    else:
        try:
            return _gl_pdf(accounts, bs_ids, begin, end, dr_cr_filter, company)
        except Exception as e:
            flash(f'PDF error: {e}. Install reportlab: pip install reportlab', 'error')
            return redirect(url_for('reports_page'))


@app.route('/reports/formatted')
def report_formatted():
    """Generate formatted BS or IS report. Placeholder — redirects to print view."""
    report_name = request.args.get('report', 'BS')
    end = request.args.get('end', '')
    begin = request.args.get('begin', '')
    fmt = request.args.get('fmt', 'pdf')
    
    reports = models.get_reports()
    rpt = next((r for r in reports if r['name'] == report_name), None)
    if not rpt:
        flash(f'Report "{report_name}" not found', 'error')
        return redirect(url_for('reports_page'))
    
    # Redirect to existing print view with parameters
    params = f'begin={begin}&end={end}' if begin else f'end={end}'
    return redirect(f'/report/{rpt["id"]}/print?{params}')


# ─── Setup Subledgers ──────────────────────────────────────────────

@app.route('/api/setup-detailed-ar', methods=['POST'])
def api_setup_detailed_ar():
    """One-click scaffold for Detailed AR subledger report."""
    if not models.get_db_path():
        return jsonify({'ok': False, 'error': 'No books open'})
    try:
        result = models.setup_detailed_ar()
        return jsonify({'ok': True, 'message': result})
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/setup-detailed-ap', methods=['POST'])
def api_setup_detailed_ap():
    """One-click scaffold for Detailed AP subledger report."""
    if not models.get_db_path():
        return jsonify({'ok': False, 'error': 'No books open'})
    try:
        result = models.setup_detailed_ap()
        return jsonify({'ok': True, 'message': result})
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ─── Entry Point ────────────────────────────────────────────────────

def main():
    import webbrowser
    
    cfg = load_config()
    
    # If we have a last-opened file, open it automatically
    last = cfg.get('last_opened', '')
    if last and os.path.exists(last):
        models.set_db_path(last)
        models._ensure_columns()  # Migrate older DBs
        company = models.get_meta('company_name', 'My Books')
        print(f"\n  Grid — {company}")
        print(f"  File: {last}")
    else:
        print(f"\n  Grid — Accounting")
        print(f"  No books open. Select a client from the library.")
    
    print(f"  Open http://localhost:5000 in your browser\n")
    
    # Open browser automatically
    webbrowser.open('http://localhost:5000')
    
    app.run(debug=False, port=5000)

if __name__ == '__main__':
    main()
